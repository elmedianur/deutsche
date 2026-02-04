"""Admin Shop Management"""
from aiogram import Router, F, Bot
from aiogram.types import CallbackQuery, Message
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.types import InlineKeyboardButton
from aiogram.filters import Command

from src.database import get_session
from src.database.models import FlashcardDeck, Language, Level, Day
from src.repositories import FlashcardDeckRepository
from src.core.logging import get_logger
from src.config import settings
from sqlalchemy import select

logger = get_logger(__name__)
router = Router(name="shop_admin")


def is_admin(user_id: int) -> bool:
    """Check if user is admin"""
    return user_id in settings.ADMIN_IDS or user_id in settings.SUPER_ADMIN_IDS


class ShopAdminStates(StatesGroup):
    """Shop admin states"""
    waiting_deck_name = State()
    waiting_deck_description = State()
    waiting_deck_icon = State()
    waiting_deck_price = State()
    editing_field = State()
    adding_card_front = State()
    adding_card_back = State()
    adding_card_example = State()


# ============================================================
# ADMIN SHOP MENU
# ============================================================

@router.callback_query(F.data == "admin:shop")
async def admin_shop_menu(callback: CallbackQuery):
    """Admin shop boshqaruv menusi"""
    async with get_session() as session:
        result = await session.execute(
            select(FlashcardDeck).where(FlashcardDeck.is_active == True).order_by(FlashcardDeck.id)
        )
        decks = list(result.scalars().all())
    
    text = f"""
🛒 <b>Do'kon Boshqaruvi</b>

📦 Jami decklar: {len(decks)}
"""
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="📥 Universal Import (Quiz+Flashcard)", callback_data="admin:universal_import")
    )
    builder.row(
        InlineKeyboardButton(text="➕ Yangi deck", callback_data="admin:shop:add_deck")
    )
    builder.row(
        InlineKeyboardButton(text="📦 Decklar ro'yxati", callback_data="admin:shop:decks")
    )
    builder.row(
        InlineKeyboardButton(text="📊 Sotuvlar", callback_data="admin:shop:sales")
    )
    builder.row(
        InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:panel")
    )
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


# ============================================================
# DECKS LIST
# ============================================================

@router.callback_query(F.data == "admin:shop:decks")
async def admin_decks_list(callback: CallbackQuery):
    """Decklar ro'yxati - Darajalar bo'yicha"""
    from sqlalchemy import func

    async with get_session() as session:
        # Get levels with deck counts
        result = await session.execute(
            select(Level).where(Level.is_active == True).order_by(Level.display_order)
        )
        levels = result.scalars().all()

    level_icons = {
        "A1": "🟢", "A2": "🟡", "B1": "🔵",
        "B2": "🟣", "C1": "🟠", "C2": "🔴"
    }

    text = "📦 <b>Decklar Ro'yxati</b>\n\n<i>Darajani tanlang:</i>\n"

    builder = InlineKeyboardBuilder()

    async with get_session() as session:
        for level in levels:
            deck_count_result = await session.execute(
                select(func.count(FlashcardDeck.id)).where(
                    FlashcardDeck.level_id == level.id
                )
            )
            deck_count = deck_count_result.scalar() or 0

            icon = level_icons.get(level.name.upper().split()[0], "📚")
            builder.row(InlineKeyboardButton(
                text=f"{icon} {level.name} — {deck_count} ta deck",
                callback_data=f"admin:shop:decks_level:{level.id}"
            ))

        # Decks without level
        no_level_result = await session.execute(
            select(func.count(FlashcardDeck.id)).where(
                FlashcardDeck.level_id == None
            )
        )
        no_level_count = no_level_result.scalar() or 0
        if no_level_count > 0:
            builder.row(InlineKeyboardButton(
                text=f"📂 Darajasiz — {no_level_count} ta deck",
                callback_data="admin:shop:decks_level:0"
            ))

    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:shop"))

    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.callback_query(F.data.startswith("admin:shop:decks_level:"))
async def admin_decks_by_level(callback: CallbackQuery):
    """Daraja ichidagi decklar"""
    level_id = int(callback.data.split(":")[-1])

    level_icons = {
        "A1": "🟢", "A2": "🟡", "B1": "🔵",
        "B2": "🟣", "C1": "🟠", "C2": "🔴"
    }

    async with get_session() as session:
        if level_id == 0:
            # Decks without level
            result = await session.execute(
                select(FlashcardDeck).where(
                    FlashcardDeck.level_id == None
                ).order_by(FlashcardDeck.display_order, FlashcardDeck.id)
            )
            level_name = "Darajasiz"
            icon = "📂"
        else:
            level_result = await session.execute(
                select(Level).where(Level.id == level_id)
            )
            level = level_result.scalar_one_or_none()
            if not level:
                await callback.answer("❌ Daraja topilmadi!", show_alert=True)
                return

            level_name = level.name
            icon = level_icons.get(level.name.upper().split()[0], "📚")

            result = await session.execute(
                select(FlashcardDeck).where(
                    FlashcardDeck.level_id == level_id
                ).order_by(FlashcardDeck.display_order, FlashcardDeck.id)
            )

        decks = list(result.scalars().all())

        # Get day names for decks
        day_names = {}
        for deck in decks:
            if deck.day_id:
                day_result = await session.execute(
                    select(Day).where(Day.id == deck.day_id)
                )
                day = day_result.scalar_one_or_none()
                if day:
                    day_names[deck.id] = day.display_name

    text = f"{icon} <b>{level_name} — Decklar</b>\n\n"

    builder = InlineKeyboardBuilder()

    if not decks:
        text += "<i>Bu darajada decklar yo'q.</i>\n"
    else:
        for deck in decks:
            status = "✅" if deck.is_active else "❌"
            premium = "⭐" if deck.is_premium else ""
            display = day_names.get(deck.id, deck.name)
            builder.row(InlineKeyboardButton(
                text=f"{status} {display} ({deck.cards_count}) {premium}",
                callback_data=f"admin:shop:deck:{deck.id}"
            ))

    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:shop:decks"))

    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


# ============================================================
# DECK DETAIL & EDIT
# ============================================================

@router.callback_query(F.data.startswith("admin:shop:deck:"))
async def admin_deck_detail(callback: CallbackQuery):
    """Deck tafsilotlari"""
    deck_id = int(callback.data.split(":")[-1])
    
    async with get_session() as session:
        result = await session.execute(
            select(FlashcardDeck).where(FlashcardDeck.id == deck_id)
        )
        deck = result.scalar_one_or_none()
    
    if not deck:
        await callback.answer("Deck topilmadi!", show_alert=True)
        return
    
    status = "✅ Faol" if deck.is_active else "❌ Nofaol"
    premium = "⭐ Premium" if deck.is_premium else "🆓 Bepul"
    
    text = f"""
📦 <b>{deck.icon} {deck.name}</b>

📝 Tavsif: {deck.description or "—"}
💰 Narxi: {deck.price} ⭐
📚 Kartalar: {deck.cards_count} ta
👥 O'quvchilar: {deck.users_studying}
📊 Status: {status}
💎 Turi: {premium}
"""
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="✏️ Nom", callback_data=f"admin:shop:edit:{deck_id}:name"),
        InlineKeyboardButton(text="📝 Tavsif", callback_data=f"admin:shop:edit:{deck_id}:desc")
    )
    builder.row(
        InlineKeyboardButton(text="💰 Narx", callback_data=f"admin:shop:edit:{deck_id}:price"),
        InlineKeyboardButton(text="🎨 Icon", callback_data=f"admin:shop:edit:{deck_id}:icon")
    )
    builder.row(
        InlineKeyboardButton(text="➕ Karta qo'sh", callback_data=f"admin:shop:add_card:{deck_id}")
    )
    builder.row(
        InlineKeyboardButton(text="📋 Kartalar", callback_data=f"admin:shop:cards:{deck_id}")
    )
    
    # Toggle buttons
    if deck.is_active:
        builder.row(
            InlineKeyboardButton(text="❌ Nofaol qilish", callback_data=f"admin:shop:toggle:{deck_id}:active")
        )
    else:
        builder.row(
            InlineKeyboardButton(text="✅ Faol qilish", callback_data=f"admin:shop:toggle:{deck_id}:active")
        )
    
    if deck.is_premium:
        builder.row(
            InlineKeyboardButton(text="🆓 Bepul qilish", callback_data=f"admin:shop:toggle:{deck_id}:premium")
        )
    else:
        builder.row(
            InlineKeyboardButton(text="⭐ Premium qilish", callback_data=f"admin:shop:toggle:{deck_id}:premium")
        )
    
    builder.row(
        InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"admin:shop:delete:{deck_id}")
    )
    builder.row(
        InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:shop:decks")
    )
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


# ============================================================
# ADD NEW DECK
# ============================================================

@router.callback_query(F.data == "admin:shop:add_deck")
async def admin_add_deck_start(callback: CallbackQuery, state: FSMContext):
    """Yangi deck qo'shishni boshlash"""
    await state.set_state(ShopAdminStates.waiting_deck_name)
    await state.update_data(new_deck={})
    
    await callback.message.edit_text(
        "➕ <b>Yangi Deck</b>\n\n"
        "Deck nomini kiriting:\n\n"
        "<i>Masalan: Oziq-ovqat, Kiyimlar, Hayvonlar</i>",
        reply_markup=InlineKeyboardBuilder().row(
            InlineKeyboardButton(text="❌ Bekor", callback_data="admin:shop")
        ).as_markup()
    )
    await callback.answer()


@router.message(ShopAdminStates.waiting_deck_name)
async def admin_add_deck_name(message: Message, state: FSMContext):
    """Deck nomi"""
    data = await state.get_data()
    new_deck = data.get("new_deck", {})
    new_deck["name"] = message.text.strip()
    await state.update_data(new_deck=new_deck)
    
    await state.set_state(ShopAdminStates.waiting_deck_description)
    await message.answer(
        "📝 Deck tavsifini kiriting:\n\n"
        "<i>Masalan: Kundalik hayotda ishlatiladigan oziq-ovqat so'zlari</i>"
    )


@router.message(ShopAdminStates.waiting_deck_description)
async def admin_add_deck_desc(message: Message, state: FSMContext):
    """Deck tavsifi"""
    data = await state.get_data()
    new_deck = data.get("new_deck", {})
    new_deck["description"] = message.text.strip()
    await state.update_data(new_deck=new_deck)
    
    await state.set_state(ShopAdminStates.waiting_deck_icon)
    await message.answer(
        "🎨 Deck ikonkasini kiriting (emoji):\n\n"
        "<i>Masalan: 🍎 🥗 👕 🐕</i>"
    )


@router.message(ShopAdminStates.waiting_deck_icon)
async def admin_add_deck_icon(message: Message, state: FSMContext):
    """Deck icon"""
    data = await state.get_data()
    new_deck = data.get("new_deck", {})
    new_deck["icon"] = message.text.strip()[:10]
    await state.update_data(new_deck=new_deck)
    
    await state.set_state(ShopAdminStates.waiting_deck_price)
    await message.answer(
        "💰 Deck narxini kiriting (Stars):\n\n"
        "<i>0 = Bepul\n50 = 50 Stars\n100 = 100 Stars</i>"
    )


@router.message(ShopAdminStates.waiting_deck_price)
async def admin_add_deck_price(message: Message, state: FSMContext):
    """Deck narxi va yaratish"""
    try:
        price = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Raqam kiriting!")
        return
    
    data = await state.get_data()
    new_deck = data.get("new_deck", {})
    new_deck["price"] = price
    new_deck["is_premium"] = price > 0
    
    # Database'ga saqlash
    async with get_session() as session:
        deck = FlashcardDeck(
            name=new_deck["name"],
            description=new_deck["description"],
            icon=new_deck["icon"],
            price=new_deck["price"],
            is_premium=new_deck["is_premium"],
            is_public=True,
            is_active=True,
            cards_count=0,
            users_studying=0,
            display_order=0
        )
        session.add(deck)
        await session.commit()
        deck_id = deck.id
    
    await state.clear()
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="➕ Karta qo'shish", callback_data=f"admin:shop:add_card:{deck_id}")
    )
    builder.row(
        InlineKeyboardButton(text="📦 Decklar", callback_data="admin:shop:decks")
    )
    
    await message.answer(
        f"✅ <b>Deck yaratildi!</b>\n\n"
        f"{new_deck['icon']} {new_deck['name']}\n"
        f"💰 Narx: {price} ⭐\n\n"
        f"Endi kartalar qo'shing:",
        reply_markup=builder.as_markup()
    )


# ============================================================
# EDIT DECK FIELDS
# ============================================================

@router.callback_query(F.data.startswith("admin:shop:edit:"))
async def admin_edit_field_start(callback: CallbackQuery, state: FSMContext):
    """Maydonni tahrirlash"""
    parts = callback.data.split(":")
    deck_id = int(parts[3])
    field = parts[4]
    
    await state.set_state(ShopAdminStates.editing_field)
    await state.update_data(edit_deck_id=deck_id, edit_field=field)
    
    field_names = {
        "name": "Yangi nom",
        "desc": "Yangi tavsif",
        "price": "Yangi narx (raqam)",
        "icon": "Yangi ikonka (emoji)"
    }
    
    await callback.message.edit_text(
        f"✏️ <b>{field_names.get(field, field)}</b> kiriting:",
        reply_markup=InlineKeyboardBuilder().row(
            InlineKeyboardButton(text="❌ Bekor", callback_data=f"admin:shop:deck:{deck_id}")
        ).as_markup()
    )
    await callback.answer()


@router.message(ShopAdminStates.editing_field)
async def admin_edit_field_save(message: Message, state: FSMContext):
    """Tahrirlangan maydonni saqlash"""
    data = await state.get_data()
    deck_id = data.get("edit_deck_id")
    field = data.get("edit_field")
    value = message.text.strip()
    
    async with get_session() as session:
        result = await session.execute(
            select(FlashcardDeck).where(FlashcardDeck.id == deck_id)
        )
        deck = result.scalar_one_or_none()
        
        if deck:
            if field == "name":
                deck.name = value
            elif field == "desc":
                deck.description = value
            elif field == "price":
                try:
                    deck.price = int(value)
                    deck.is_premium = deck.price > 0
                except ValueError:
                    await message.answer("❌ Raqam kiriting!")
                    return
            elif field == "icon":
                deck.icon = value[:10]
            
            await session.commit()
    
    await state.clear()
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="◀️ Deckga qaytish", callback_data=f"admin:shop:deck:{deck_id}")
    )
    
    await message.answer("✅ Saqlandi!", reply_markup=builder.as_markup())


# ============================================================
# TOGGLE & DELETE
# ============================================================

@router.callback_query(F.data.startswith("admin:shop:toggle:"))
async def admin_toggle_deck(callback: CallbackQuery):
    """Deck holatini o'zgartirish"""
    parts = callback.data.split(":")
    deck_id = int(parts[3])
    toggle_type = parts[4]
    
    async with get_session() as session:
        result = await session.execute(
            select(FlashcardDeck).where(FlashcardDeck.id == deck_id)
        )
        deck = result.scalar_one_or_none()
        
        if deck:
            if toggle_type == "active":
                deck.is_active = not deck.is_active
                msg = "✅ Faollashtirildi" if deck.is_active else "❌ Nofaol qilindi"
            elif toggle_type == "premium":
                deck.is_premium = not deck.is_premium
                msg = "⭐ Premium qilindi" if deck.is_premium else "🆓 Bepul qilindi"
            
            await session.commit()
    
    await callback.answer(msg, show_alert=True)
    
    # Refresh page
    await admin_deck_detail(callback)


@router.callback_query(F.data.startswith("admin:shop:delete:"))
async def admin_delete_deck(callback: CallbackQuery):
    """Deckni o'chirish"""
    deck_id = int(callback.data.split(":")[-1])
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="✅ Ha, o'chirish", callback_data=f"admin:shop:delete_confirm:{deck_id}"),
        InlineKeyboardButton(text="❌ Yo'q", callback_data=f"admin:shop:deck:{deck_id}")
    )
    
    await callback.message.edit_text(
        "⚠️ <b>Rostdan o'chirmoqchimisiz?</b>\n\n"
        "Bu amalni qaytarib bo'lmaydi!",
        reply_markup=builder.as_markup()
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin:shop:delete_confirm:"))
async def admin_delete_confirm(callback: CallbackQuery):
    """O'chirishni tasdiqlash"""
    deck_id = int(callback.data.split(":")[-1])
    
    async with get_session() as session:
        result = await session.execute(
            select(FlashcardDeck).where(FlashcardDeck.id == deck_id)
        )
        deck = result.scalar_one_or_none()
        if deck:
            await session.delete(deck)
            await session.commit()
    
    await callback.answer("🗑 O'chirildi!", show_alert=True)
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="◀️ Decklar", callback_data="admin:shop:decks")
    )
    
    await callback.message.edit_text(
        "✅ Deck o'chirildi!",
        reply_markup=builder.as_markup()
    )


# ============================================================
# ADD CARDS TO DECK
# ============================================================

@router.callback_query(F.data.startswith("admin:shop:add_card:"))
async def admin_add_card_start(callback: CallbackQuery, state: FSMContext):
    """Deckga karta qo'shish"""
    deck_id = int(callback.data.split(":")[-1])
    
    await state.set_state(ShopAdminStates.adding_card_front)
    await state.update_data(add_card_deck_id=deck_id, new_card={})
    
    await callback.message.edit_text(
        "➕ <b>Yangi Karta</b>\n\n"
        "So'zni kiriting (nemischa):\n\n"
        "<i>Masalan: die Milch</i>",
        reply_markup=InlineKeyboardBuilder().row(
            InlineKeyboardButton(text="❌ Bekor", callback_data=f"admin:shop:deck:{deck_id}")
        ).as_markup()
    )
    await callback.answer()


@router.message(ShopAdminStates.adding_card_front)
async def admin_add_card_front(message: Message, state: FSMContext):
    """Karta front"""
    data = await state.get_data()
    new_card = data.get("new_card", {})
    new_card["front"] = message.text.strip()
    await state.update_data(new_card=new_card)
    
    await state.set_state(ShopAdminStates.adding_card_back)
    await message.answer(
        "Tarjimasini kiriting (o'zbekcha):\n\n"
        "<i>Masalan: Sut</i>"
    )


@router.message(ShopAdminStates.adding_card_back)
async def admin_add_card_back(message: Message, state: FSMContext):
    """Karta back"""
    data = await state.get_data()
    new_card = data.get("new_card", {})
    new_card["back"] = message.text.strip()
    await state.update_data(new_card=new_card)
    
    await state.set_state(ShopAdminStates.adding_card_example)
    await message.answer(
        "Misol gap kiriting:\n\n"
        "<i>Masalan: Ich trinke Milch.</i>\n\n"
        "Yoki /skip bosing"
    )


@router.message(ShopAdminStates.adding_card_example)
async def admin_add_card_example(message: Message, state: FSMContext):
    """Karta example va saqlash"""
    data = await state.get_data()
    new_card = data.get("new_card", {})
    deck_id = data.get("add_card_deck_id")
    
    example = "" if message.text == "/skip" else message.text.strip()
    
    from src.database.models import Flashcard
    
    async with get_session() as session:
        # Kartani qo'shish
        card = Flashcard(
            deck_id=deck_id,
            front_text=new_card["front"],
            back_text=new_card["back"],
            example_sentence=example,
            times_shown=0,
            times_known=0,
            display_order=0,
            is_active=True
        )
        session.add(card)
        
        # Deck cards_count yangilash
        result = await session.execute(
            select(FlashcardDeck).where(FlashcardDeck.id == deck_id)
        )
        deck = result.scalar_one_or_none()
        if deck:
            deck.cards_count += 1
        
        await session.commit()
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="➕ Yana qo'shish", callback_data=f"admin:shop:add_card:{deck_id}")
    )
    builder.row(
        InlineKeyboardButton(text="◀️ Deckga qaytish", callback_data=f"admin:shop:deck:{deck_id}")
    )
    
    await state.clear()
    await message.answer(
        f"✅ <b>Karta qo'shildi!</b>\n\n"
        f"🔤 {new_card['front']}\n"
        f"📖 {new_card['back']}",
        reply_markup=builder.as_markup()
    )


# ============================================================
# VIEW CARDS
# ============================================================

@router.callback_query(F.data.startswith("admin:shop:cards:"))
async def admin_view_cards(callback: CallbackQuery):
    """Deck kartalari ro'yxati"""
    deck_id = int(callback.data.split(":")[-1])
    
    from src.database.models import Flashcard
    
    async with get_session() as session:
        result = await session.execute(
            select(Flashcard).where(Flashcard.deck_id == deck_id).limit(50)
        )
        cards = list(result.scalars().all())
    
    if not cards:
        await callback.answer("Kartalar yo'q!", show_alert=True)
        return
    
    text = f"📋 <b>Kartalar</b> ({len(cards)} ta)\n\n"
    
    for i, card in enumerate(cards[:20], 1):
        text += f"{i}. {card.front_text} — {card.back_text}\n"
    
    if len(cards) > 20:
        text += f"\n... va yana {len(cards) - 20} ta"
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="➕ Karta qo'shish", callback_data=f"admin:shop:add_card:{deck_id}")
    )
    builder.row(
        InlineKeyboardButton(text="◀️ Orqaga", callback_data=f"admin:shop:deck:{deck_id}")
    )
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


# ============================================================
# SALES STATISTICS
# ============================================================

@router.callback_query(F.data == "admin:shop:sales")
async def admin_sales_stats(callback: CallbackQuery):
    """Sotuvlar statistikasi"""
    from src.database.models import UserDeckPurchase
    from sqlalchemy import func
    
    async with get_session() as session:
        # Jami sotuvlar
        result = await session.execute(
            select(func.count(UserDeckPurchase.id), func.sum(UserDeckPurchase.price_paid))
        )
        row = result.one()
        total_sales = row[0] or 0
        total_revenue = row[1] or 0
        
        # Deck bo'yicha
        result = await session.execute(
            select(
                FlashcardDeck.name,
                func.count(UserDeckPurchase.id)
            ).join(
                UserDeckPurchase, UserDeckPurchase.deck_id == FlashcardDeck.id
            ).group_by(FlashcardDeck.id).order_by(func.count(UserDeckPurchase.id).desc()).limit(10)
        )
        top_decks = list(result.all())
    
    text = f"""
📊 <b>Sotuvlar Statistikasi</b>

💰 Jami sotuvlar: {total_sales} ta
⭐ Jami daromad: {total_revenue} Stars

<b>Top decklar:</b>
"""
    
    for i, (name, count) in enumerate(top_decks, 1):
        text += f"{i}. {name}: {count} ta\n"
    
    if not top_decks:
        text += "<i>Hali sotuvlar yo'q</i>"
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:shop")
    )

    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


# ============================================================
# MARKET CONTENT IMPORT - Darajalar va Mavzular
# ============================================================

@router.callback_query(F.data == "admin:market_import")
async def market_import_menu(callback: CallbackQuery):
    """Market kontent import menusi"""
    if not is_admin(callback.from_user.id):
        return

    text = """
📥 <b>Market Kontent Import</b>

Excel faylda quyidagi ustunlar bo'lishi kerak:

<b>Majburiy ustunlar:</b>
• <code>level</code> - Daraja (A1, A2, B1, B2, C1, C2)
• <code>order</code> - Tartib raqami (1, 2, 3...)
• <code>name</code> - Mavzu nomi (O'zbekcha)
• <code>topic</code> - Mavzu (Nemischa)

<b>Ixtiyoriy ustunlar:</b>
• <code>description</code> - Tavsif
• <code>is_premium</code> - Pulli (1) yoki Bepul (0)
• <code>price</code> - Narx (Stars)

<b>Namuna:</b>
| level | order | name | topic | is_premium | price |
|-------|-------|------|-------|------------|-------|
| A1 | 1 | Alfavit | Das Alphabet | 0 | 0 |
| A1 | 2 | Salomlashish | Begrüßung | 0 | 0 |
| A1 | 3 | Tanishuv | Vorstellung | 0 | 0 |
| A1 | 4 | Oila | Familie | 1 | 50 |
| A2 | 1 | Kuchli fe'llar | Starke Verben | 1 | 75 |

Faylni yuboring va <code>/import_market</code> buyrug'ini yozing.
"""

    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="📄 Namuna yuklab olish", callback_data="admin:market_template"))
    builder.row(InlineKeyboardButton(text="📋 Mavjud mavzular", callback_data="admin:market_topics"))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:shop"))

    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.callback_query(F.data == "admin:market_template")
async def download_market_template(callback: CallbackQuery):
    """Market Excel template yuklab olish"""
    if not is_admin(callback.from_user.id):
        return

    await callback.answer("📄 Template tayyorlanmoqda...")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import tempfile
        import os

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Market Topics"

        headers = ["level", "order", "name", "topic", "description", "is_premium", "price"]
        descriptions = ["Daraja", "Tartib", "Nomi (UZ)", "Nomi (DE)", "Tavsif", "Pulli (1/0)", "Narx (Stars)"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        free_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        premium_fill = PatternFill(start_color="FFECB3", end_color="FFECB3", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Headers
        for col, (h, d) in enumerate(zip(headers, descriptions), 1):
            ws.cell(row=1, column=col, value=h).font = header_font
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).border = border
            ws.cell(row=2, column=col, value=d).font = Font(italic=True, color="666666")
            ws.cell(row=2, column=col).border = border

        # A1 topics (namuna)
        a1_topics = [
            ["A1", 1, "🔤 Alfavit va talaffuz", "Das Alphabet", "Nemis alifbosi", 0, 0],
            ["A1", 2, "👋 Salomlashish", "Begrüßung", "Salom va xayr", 0, 0],
            ["A1", 3, "🤝 O'zini tanishtirish", "Vorstellung", "Tanishish", 0, 0],
            ["A1", 4, "🔢 Raqamlar", "Zahlen", "1-100 raqamlar", 1, 30],
            ["A1", 5, "👨‍👩‍👧‍👦 Oila", "Familie", "Oila a'zolari", 1, 30],
            ["A1", 6, "🧍 Tana a'zolari", "Körperteile", "Inson tanasi", 1, 40],
            ["A1", 7, "🎨 Ranglar", "Farben", "Asosiy ranglar", 1, 25],
            ["A1", 8, "📅 Kunlar va oylar", "Tage und Monate", "Kalendar", 1, 35],
            ["A1", 9, "🕐 Vaqt", "Uhrzeit", "Soat aytish", 1, 35],
            ["A1", 10, "🍽️ Ovqatlanish", "Essen und Trinken", "Taomlar", 1, 40],
        ]

        a2_topics = [
            ["A2", 1, "💪 Modal fe'llar", "Modalverben", "können, müssen...", 1, 50],
            ["A2", 2, "📍 Predloglar", "Präpositionen", "in, an, auf...", 1, 50],
            ["A2", 3, "✨ Sifatlar", "Adjektive", "Sifat darajalari", 1, 50],
            ["A2", 4, "✈️ Sayohat", "Reisen", "Mehmonxona, transport", 1, 60],
            ["A2", 5, "🏥 Sog'liq", "Gesundheit", "Shifokor, kasallik", 1, 60],
        ]

        b1_topics = [
            ["B1", 1, "🔄 Majhul nisbat", "Passiv", "Grammatika", 1, 75],
            ["B1", 2, "🔗 Ergash gaplar", "Relativsätze", "Murakkab gaplar", 1, 75],
            ["B1", 3, "💼 Ish va martaba", "Arbeit und Karriere", "Professional til", 1, 80],
        ]

        all_topics = a1_topics + a2_topics + b1_topics

        for r, row_data in enumerate(all_topics, 3):
            is_premium = row_data[5]
            fill = premium_fill if is_premium else free_fill
            for c, v in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = border
                cell.fill = fill

        # Column widths
        for col, w in enumerate([8, 8, 30, 25, 30, 12, 10], 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        file_path = os.path.join(tempfile.gettempdir(), "market_template.xlsx")
        wb.save(file_path)

        from aiogram.types import FSInputFile
        await callback.message.answer_document(
            document=FSInputFile(file_path, filename="market_template.xlsx"),
            caption="📄 <b>Market Template</b>\n\n"
                   "🟢 Yashil = Bepul mavzular\n"
                   "🟡 Sariq = Pulli mavzular\n\n"
                   "To'ldiring va <code>/import_market</code> bilan yuklang."
        )
        os.remove(file_path)

    except ImportError:
        await callback.message.answer("❌ <code>pip install openpyxl</code> kerak!")
    except Exception as e:
        await callback.message.answer(f"❌ Xatolik: {e}")


@router.callback_query(F.data == "admin:market_topics")
async def show_market_topics(callback: CallbackQuery):
    """Mavjud mavzularni ko'rsatish"""
    if not is_admin(callback.from_user.id):
        return

    async with get_session() as session:
        # Get all levels with their days
        result = await session.execute(
            select(Level).order_by(Level.display_order)
        )
        levels = result.scalars().all()

    text = "📋 <b>Mavjud Mavzular (Market):</b>\n\n"

    for level in levels:
        text += f"<b>{level.name}</b>:\n"
        # Get days for this level
        async with get_session() as session:
            result = await session.execute(
                select(Day).where(Day.level_id == level.id).order_by(Day.day_number)
            )
            days = result.scalars().all()

        free_count = sum(1 for d in days if not d.is_premium)
        premium_count = sum(1 for d in days if d.is_premium)

        text += f"   🆓 Bepul: {free_count} | ⭐ Pulli: {premium_count}\n"

        for day in days[:5]:  # First 5 only
            status = "⭐" if day.is_premium else "🆓"
            price = f" ({day.price}💫)" if day.is_premium and hasattr(day, 'price') and day.price else ""
            text += f"   {status} {day.name}{price}\n"

        if len(days) > 5:
            text += f"   <i>...va yana {len(days) - 5} ta</i>\n"
        text += "\n"

    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="📥 Import qilish", callback_data="admin:market_import"))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:shop"))

    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.message(Command("import_market"))
async def import_market_command(message: Message, state: FSMContext, bot: Bot):
    """Market kontentni Excel'dan import qilish"""
    if not is_admin(message.from_user.id):
        return

    data = await state.get_data()
    file_id = data.get("pending_excel_file_id")

    if not file_id:
        await message.answer(
            "❌ Avval Excel faylni yuboring!\n\n"
            "1. Excel faylni yuboring\n"
            "2. <code>/import_market</code> buyrug'ini yozing"
        )
        return

    await message.answer("⏳ Market kontent import qilinmoqda...")

    try:
        import openpyxl
        import tempfile
        import os

        file = await bot.get_file(file_id)
        file_path = os.path.join(tempfile.gettempdir(), f"market_{message.from_user.id}.xlsx")
        await bot.download_file(file.file_path, file_path)

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        required = ['level', 'order', 'name']
        missing = [c for c in required if c not in headers]

        if missing:
            await message.answer(f"❌ Ustunlar yo'q: {', '.join(missing)}")
            os.remove(file_path)
            return

        # Parse data
        topics = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            level_name = str(row_data.get('level') or '').strip().upper()
            if not level_name or level_name in ['DARAJA', 'LEVEL']:
                continue
            topics.append(row_data)

        if not topics:
            await message.answer("❌ Ma'lumotlar topilmadi!")
            os.remove(file_path)
            return

        # Get or create German language
        async with get_session() as session:
            result = await session.execute(select(Language).where(Language.code == "de"))
            german = result.scalar_one_or_none()

            if not german:
                german = Language(
                    name="Nemis tili",
                    code="de",
                    flag="🇩🇪",
                    is_active=True,
                    display_order=1
                )
                session.add(german)
                await session.flush()

            # Level descriptions
            level_info = {
                "A1": ("Boshlang'ich", "Elementar daraja", 1, False),
                "A2": ("Asosiy", "Asosiy til bilimi", 2, False),
                "B1": ("O'rta", "Mustaqil daraja", 3, True),
                "B2": ("O'rta-yuqori", "Yuqori o'rta", 4, True),
                "C1": ("Ilg'or", "Professional daraja", 5, True),
                "C2": ("Mahorat", "Ona tili darajasi", 6, True),
            }

            # Process topics
            levels_created = 0
            topics_created = 0
            topics_updated = 0

            # Group by level
            by_level = {}
            for t in topics:
                level_name = str(t.get('level') or '').strip().upper()
                if level_name not in by_level:
                    by_level[level_name] = []
                by_level[level_name].append(t)

            for level_name, level_topics in by_level.items():
                # Get or create level
                result = await session.execute(
                    select(Level).where(
                        Level.language_id == german.id,
                        Level.name == level_name
                    )
                )
                level = result.scalar_one_or_none()

                if not level:
                    info = level_info.get(level_name, ("Daraja", "Daraja", 99, True))
                    level = Level(
                        language_id=german.id,
                        name=level_name,
                        description=f"{info[0]} - {info[1]}",
                        is_premium=info[3],
                        is_active=True,
                        display_order=info[2]
                    )
                    session.add(level)
                    await session.flush()
                    levels_created += 1

                # Process topics for this level
                for t in level_topics:
                    order = int(t.get('order') or 1)
                    name = str(t.get('name') or '').strip()
                    topic_de = str(t.get('topic') or '').strip()
                    description = str(t.get('description') or '').strip() or None
                    is_premium = bool(int(t.get('is_premium') or 0))
                    price = int(t.get('price') or 0)

                    if not name:
                        continue

                    # Check if day exists
                    result = await session.execute(
                        select(Day).where(
                            Day.level_id == level.id,
                            Day.day_number == order
                        )
                    )
                    day = result.scalar_one_or_none()

                    if day:
                        # Update existing
                        day.name = name
                        day.topic = topic_de or None
                        day.description = description
                        day.is_premium = is_premium
                        day.price = price
                        topics_updated += 1
                    else:
                        # Create new
                        day = Day(
                            level_id=level.id,
                            day_number=order,
                            name=name,
                            topic=topic_de or None,
                            description=description,
                            is_premium=is_premium,
                            price=price,
                            is_active=True
                        )
                        session.add(day)
                        topics_created += 1

            await session.commit()

        os.remove(file_path)
        await state.update_data(pending_excel_file_id=None)

        await message.answer(
            f"✅ <b>Market import yakunlandi!</b>\n\n"
            f"📊 Darajalar: {levels_created} ta yangi\n"
            f"📚 Mavzular: {topics_created} ta yangi, {topics_updated} ta yangilandi\n\n"
            f"<i>Jami: {topics_created + topics_updated} ta mavzu</i>"
        )

    except ImportError:
        await message.answer("❌ <code>pip install openpyxl</code> kerak!")
    except Exception as e:
        logger.error(f"Market import error: {e}")
        await message.answer(f"❌ Xatolik: {e}")


# ============================================================
# UNIVERSAL IMPORT - Bitta so'z -> Quiz + Flashcard
# ============================================================

@router.callback_query(F.data == "admin:universal_import")
async def universal_import_menu(callback: CallbackQuery):
    """Universal kontent import menusi"""
    if not is_admin(callback.from_user.id):
        return

    text = """
📥 <b>Universal Import</b>
<i>Bitta so'z → Quiz + Flashcard</i>

Excel faylda quyidagi ustunlar bo'lishi kerak:

<b>Majburiy ustunlar:</b>
• <code>level</code> - Daraja (A1, A2, B1...)
• <code>topic_uz</code> - Mavzu nomi (O'zbekcha)
• <code>german</code> - Nemischa so'z
• <code>correct</code> - To'g'ri javob (O'zbekcha)
• <code>wrong1</code> - Noto'g'ri variant 1
• <code>wrong2</code> - Noto'g'ri variant 2
• <code>wrong3</code> - Noto'g'ri variant 3

<b>Ixtiyoriy ustunlar:</b>
• <code>order</code> - Mavzu tartib raqami
• <code>topic_de</code> - Mavzu (Nemischa)
• <code>example_de</code> - Misol gap (Nemischa)
• <code>example_uz</code> - Misol tarjimasi
• <code>is_premium</code> - Pulli (1/0)
• <code>price</code> - Narx (Stars)

<b>Namuna:</b>
<code>
| level | topic_uz | german | correct | wrong1 | wrong2 | wrong3 |
| A1 | Salomlashish | Guten Tag | Xayrli kun | Xayr | Salom | Rahmat |
</code>

1. Excel faylni yuboring
2. <code>/universal_import</code> buyrug'ini yozing
"""

    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="📄 Namuna yuklab olish", callback_data="admin:universal_template"))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:shop"))

    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.callback_query(F.data == "admin:universal_template")
async def download_universal_template(callback: CallbackQuery):
    """Universal Excel template yuklab olish"""
    if not is_admin(callback.from_user.id):
        return

    await callback.answer("📄 Template tayyorlanmoqda...")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        import tempfile
        import os

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Universal Import"

        headers = ["level", "order", "topic_uz", "topic_de", "german", "correct", "wrong1", "wrong2", "wrong3", "example_de", "example_uz", "is_premium", "price"]
        descriptions = ["Daraja", "Tartib", "Mavzu (UZ)", "Mavzu (DE)", "Nemischa", "To'g'ri javob", "Noto'g'ri 1", "Noto'g'ri 2", "Noto'g'ri 3", "Misol (DE)", "Misol (UZ)", "Pulli (1/0)", "Narx"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        free_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        premium_fill = PatternFill(start_color="FFECB3", end_color="FFECB3", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Headers
        for col, (h, d) in enumerate(zip(headers, descriptions), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

            desc_cell = ws.cell(row=2, column=col, value=d)
            desc_cell.font = Font(italic=True, color="666666")
            desc_cell.border = border

        # Sample data - A1 Salomlashish (bepul)
        # level, order, topic_uz, topic_de, german, correct, wrong1, wrong2, wrong3, example_de, example_uz, is_premium, price
        sample_data = [
            ["A1", 1, "Salomlashish", "Begrüßung", "Guten Morgen", "Xayrli tong", "Xayrli kun", "Xayrli oqshom", "Salom", "Guten Morgen, Herr Schmidt!", "Xayrli tong, janob Shmidt!", 0, 0],
            ["A1", 1, "Salomlashish", "Begrüßung", "Guten Tag", "Xayrli kun", "Xayrli tong", "Salom", "Xayr", "Guten Tag, Frau Müller!", "Xayrli kun, xonim Myuller!", 0, 0],
            ["A1", 1, "Salomlashish", "Begrüßung", "Guten Abend", "Xayrli oqshom", "Xayrli kun", "Xayrli tong", "Ko'rishguncha", "Guten Abend zusammen!", "Hammaga xayrli oqshom!", 0, 0],
            ["A1", 1, "Salomlashish", "Begrüßung", "Hallo", "Salom", "Xayr", "Rahmat", "Kechirasiz", "Hallo, wie geht's?", "Salom, qalaysiz?", 0, 0],
            ["A1", 1, "Salomlashish", "Begrüßung", "Tschüss", "Xayr", "Salom", "Rahmat", "Kechirasiz", "Tschüss, bis morgen!", "Xayr, ertaga ko'rishguncha!", 0, 0],
            ["A1", 1, "Salomlashish", "Begrüßung", "Auf Wiedersehen", "Ko'rishguncha", "Salom", "Xayr", "Rahmat", "Auf Wiedersehen und danke!", "Ko'rishguncha va rahmat!", 0, 0],

            # A1 Raqamlar (pulli)
            ["A1", 2, "Raqamlar", "Zahlen", "eins", "bir", "ikki", "uch", "to'rt", "Ich habe eins.", "Menda bitta bor.", 1, 30],
            ["A1", 2, "Raqamlar", "Zahlen", "zwei", "ikki", "bir", "uch", "besh", "Zwei Kaffee, bitte.", "Ikkita qahva, iltimos.", 1, 30],
            ["A1", 2, "Raqamlar", "Zahlen", "drei", "uch", "ikki", "to'rt", "olti", "Drei Äpfel kosten 2 Euro.", "Uchta olma 2 yevro turadi.", 1, 30],
            ["A1", 2, "Raqamlar", "Zahlen", "vier", "to'rt", "uch", "besh", "ikki", "Ich habe vier Geschwister.", "Mening to'rtta aka-ukam bor.", 1, 30],
            ["A1", 2, "Raqamlar", "Zahlen", "fünf", "besh", "to'rt", "olti", "uch", "Es ist fünf Uhr.", "Soat besh.", 1, 30],

            # A1 Oila (pulli)
            ["A1", 3, "Oila", "Familie", "die Mutter", "ona", "ota", "aka", "opa", "Meine Mutter heißt Anna.", "Onamning ismi Anna.", 1, 40],
            ["A1", 3, "Oila", "Familie", "der Vater", "ota", "ona", "uka", "buvi", "Mein Vater arbeitet viel.", "Otam ko'p ishlaydi.", 1, 40],
            ["A1", 3, "Oila", "Familie", "die Schwester", "opa/singil", "aka", "ona", "buvi", "Meine Schwester ist 20 Jahre alt.", "Opam 20 yoshda.", 1, 40],
            ["A1", 3, "Oila", "Familie", "der Bruder", "aka/uka", "opa", "ota", "ona", "Mein Bruder studiert Medizin.", "Akam tibbiyot o'qiydi.", 1, 40],

            # A2 Sifatlar (pulli)
            ["A2", 1, "Sifatlar", "Adjektive", "groß", "katta", "kichik", "tez", "sekin", "Das Haus ist sehr groß.", "Uy juda katta.", 1, 50],
            ["A2", 1, "Sifatlar", "Adjektive", "klein", "kichik", "katta", "uzun", "qisqa", "Die Katze ist klein.", "Mushuk kichik.", 1, 50],
            ["A2", 1, "Sifatlar", "Adjektive", "schön", "chiroyli", "xunuk", "katta", "kichik", "Das Wetter ist schön.", "Ob-havo chiroyli.", 1, 50],
            ["A2", 1, "Sifatlar", "Adjektive", "schnell", "tez", "sekin", "katta", "kichik", "Der Zug ist schnell.", "Poyezd tez.", 1, 50],
        ]

        for r, row_data in enumerate(sample_data, 3):
            is_premium = row_data[11]  # is_premium is now at index 11
            fill = premium_fill if is_premium else free_fill
            for c, v in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = border
                cell.fill = fill

        # Column widths (13 columns now)
        widths = [8, 8, 16, 16, 20, 18, 15, 15, 15, 30, 30, 10, 8]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        file_path = os.path.join(tempfile.gettempdir(), "universal_template.xlsx")
        wb.save(file_path)

        from aiogram.types import FSInputFile
        await callback.message.answer_document(
            document=FSInputFile(file_path, filename="universal_template.xlsx"),
            caption="📄 <b>Universal Import Template</b>\n\n"
                   "🟢 Yashil = Bepul mavzular\n"
                   "🟡 Sariq = Pulli mavzular\n\n"
                   "Har bir qator:\n"
                   "• Quiz savoliga aylanadi\n"
                   "• Flashcard kartaga aylanadi\n\n"
                   "To'ldiring va <code>/universal_import</code> bilan yuklang."
        )
        os.remove(file_path)

    except ImportError:
        await callback.message.answer("❌ <code>pip install openpyxl</code> kerak!")
    except Exception as e:
        await callback.message.answer(f"❌ Xatolik: {e}")


@router.message(Command("universal_import"))
async def universal_import_command(message: Message, state: FSMContext, bot: Bot):
    """Universal import - bitta so'z Quiz + Flashcard ga"""
    if not is_admin(message.from_user.id):
        return

    data = await state.get_data()
    file_id = data.get("pending_excel_file_id")

    if not file_id:
        await message.answer(
            "❌ Avval Excel faylni yuboring!\n\n"
            "1. Excel faylni yuboring\n"
            "2. <code>/universal_import</code> buyrug'ini yozing"
        )
        return

    await message.answer("⏳ Universal import qilinmoqda...\nBu biroz vaqt olishi mumkin.")

    try:
        import openpyxl
        import tempfile
        import os
        import random

        file = await bot.get_file(file_id)
        file_path = os.path.join(tempfile.gettempdir(), f"universal_{message.from_user.id}.xlsx")
        await bot.download_file(file.file_path, file_path)

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
        required = ['level', 'topic_uz', 'german', 'correct', 'wrong1', 'wrong2', 'wrong3']
        missing = [c for c in required if c not in headers]

        if missing:
            await message.answer(f"❌ Ustunlar yo'q: {', '.join(missing)}")
            os.remove(file_path)
            return

        # Parse all rows
        rows_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            level_name = str(row_dict.get('level') or '').strip().upper()
            german = str(row_dict.get('german') or '').strip()
            correct = str(row_dict.get('correct') or '').strip()

            if not level_name or not german or not correct:
                continue
            if level_name in ['DARAJA', 'LEVEL']:
                continue

            rows_data.append(row_dict)

        if not rows_data:
            await message.answer("❌ Ma'lumotlar topilmadi!")
            os.remove(file_path)
            return

        from src.database.models import Question, Flashcard, FlashcardDeck

        async with get_session() as session:
            # Get or create German language
            result = await session.execute(select(Language).where(Language.code == "de"))
            german_lang = result.scalar_one_or_none()

            if not german_lang:
                german_lang = Language(
                    name="Nemis tili",
                    code="de",
                    flag="🇩🇪",
                    is_active=True,
                    display_order=1
                )
                session.add(german_lang)
                await session.flush()

            # Level info
            level_info = {
                "A1": ("Boshlang'ich", 1, False),
                "A2": ("Asosiy", 2, False),
                "B1": ("O'rta", 3, True),
                "B2": ("O'rta-yuqori", 4, True),
                "C1": ("Ilg'or", 5, True),
                "C2": ("Mahorat", 6, True),
            }

            # Group by level and topic
            by_level_topic = {}
            for row in rows_data:
                level_name = str(row.get('level') or '').strip().upper()
                topic_uz = str(row.get('topic_uz') or '').strip()
                key = (level_name, topic_uz)
                if key not in by_level_topic:
                    by_level_topic[key] = []
                by_level_topic[key].append(row)

            stats = {
                "levels": 0,
                "topics": 0,
                "questions": 0,
                "flashcards": 0,
                "decks": 0
            }

            # Process each level-topic group
            for (level_name, topic_uz), words in by_level_topic.items():
                if not words:
                    continue

                first_word = words[0]
                order = int(first_word.get('order') or 1)
                topic_de = str(first_word.get('topic_de') or '').strip() or None
                is_premium = bool(int(first_word.get('is_premium') or 0))
                price = int(first_word.get('price') or 0)

                # Get or create Level
                result = await session.execute(
                    select(Level).where(
                        Level.language_id == german_lang.id,
                        Level.name == level_name
                    )
                )
                level = result.scalar_one_or_none()

                if not level:
                    info = level_info.get(level_name, ("Daraja", 99, True))
                    level = Level(
                        language_id=german_lang.id,
                        name=level_name,
                        description=info[0],
                        is_premium=info[2],
                        is_active=True,
                        display_order=info[1]
                    )
                    session.add(level)
                    await session.flush()
                    stats["levels"] += 1

                # Get or create Day/Topic
                result = await session.execute(
                    select(Day).where(
                        Day.level_id == level.id,
                        Day.name == topic_uz
                    )
                )
                day = result.scalar_one_or_none()

                if not day:
                    # Find max day_number for this level
                    result = await session.execute(
                        select(Day).where(Day.level_id == level.id).order_by(Day.day_number.desc()).limit(1)
                    )
                    last_day = result.scalar_one_or_none()
                    next_day_number = (last_day.day_number + 1) if last_day else order

                    day = Day(
                        level_id=level.id,
                        day_number=next_day_number,
                        name=topic_uz,
                        topic=topic_de,
                        is_premium=is_premium,
                        price=price,
                        is_active=True
                    )
                    session.add(day)
                    await session.flush()
                    stats["topics"] += 1

                # Get or create Flashcard Deck for this topic (linked to Day)
                deck_name = f"{level_name} - {topic_uz}"
                result = await session.execute(
                    select(FlashcardDeck).where(FlashcardDeck.name == deck_name)
                )
                deck = result.scalar_one_or_none()

                if not deck:
                    deck = FlashcardDeck(
                        name=deck_name,
                        description=f"{level_name} darajasi - {topic_uz} mavzusi so'zlari",
                        icon="📚",
                        price=price,
                        is_premium=is_premium,
                        is_public=True,
                        is_active=True,
                        cards_count=0,
                        users_studying=0,
                        display_order=0,
                        day_id=day.id  # Link to Day for unified purchase
                    )
                    session.add(deck)
                    await session.flush()
                    stats["decks"] += 1
                else:
                    # Update existing deck to link to day
                    deck.day_id = day.id

                # Process each word
                for word in words:
                    german_word = str(word.get('german') or '').strip()
                    correct_answer = str(word.get('correct') or '').strip()
                    wrong1 = str(word.get('wrong1') or '').strip()
                    wrong2 = str(word.get('wrong2') or '').strip()
                    wrong3 = str(word.get('wrong3') or '').strip()
                    example_de = str(word.get('example_de') or '').strip() or None
                    example_uz = str(word.get('example_uz') or '').strip() or None

                    if not german_word or not correct_answer:
                        continue

                    # Shuffle options for quiz (randomize position of correct answer)
                    import random
                    all_options = [correct_answer, wrong1, wrong2, wrong3]
                    random.shuffle(all_options)

                    # Find correct option letter (A, B, C, D)
                    correct_index = all_options.index(correct_answer)
                    correct_option = ['A', 'B', 'C', 'D'][correct_index]

                    # Build explanation
                    explanation = f"✅ To'g'ri javob: {correct_answer}"
                    if example_de and example_uz:
                        explanation += f"\n\n📝 Misol:\n{example_de}\n({example_uz})"

                    # === Create Quiz Question ===
                    # Check if question exists
                    result = await session.execute(
                        select(Question).where(
                            Question.day_id == day.id,
                            Question.question_text == german_word
                        )
                    )
                    existing_q = result.scalar_one_or_none()

                    if not existing_q:
                        question = Question(
                            day_id=day.id,
                            question_text=german_word,
                            option_a=all_options[0],
                            option_b=all_options[1],
                            option_c=all_options[2],
                            option_d=all_options[3],
                            correct_option=correct_option,
                            explanation=explanation,
                            difficulty=1,
                            is_active=True
                        )
                        session.add(question)
                        stats["questions"] += 1

                    # === Create Flashcard ===
                    # Check if flashcard exists
                    result = await session.execute(
                        select(Flashcard).where(
                            Flashcard.deck_id == deck.id,
                            Flashcard.front_text == german_word
                        )
                    )
                    existing_fc = result.scalar_one_or_none()

                    if not existing_fc:
                        flashcard = Flashcard(
                            deck_id=deck.id,
                            front_text=german_word,
                            back_text=correct_answer,
                            example_sentence=f"{example_de}\n({example_uz})" if example_de else None,
                            times_shown=0,
                            times_known=0,
                            display_order=0,
                            is_active=True
                        )
                        session.add(flashcard)
                        stats["flashcards"] += 1

                # Update deck cards count
                result = await session.execute(
                    select(Flashcard).where(Flashcard.deck_id == deck.id)
                )
                all_cards = result.scalars().all()
                deck.cards_count = len(all_cards)

            await session.commit()

        os.remove(file_path)
        await state.update_data(pending_excel_file_id=None)

        await message.answer(
            f"✅ <b>Universal import yakunlandi!</b>\n\n"
            f"📊 Darajalar: {stats['levels']} ta yangi\n"
            f"📁 Mavzular: {stats['topics']} ta yangi\n"
            f"📦 Decklar: {stats['decks']} ta yangi\n"
            f"❓ Quiz savollari: {stats['questions']} ta\n"
            f"🃏 Flashcard kartalar: {stats['flashcards']} ta\n\n"
            f"<i>Har bir so'z Quiz va Flashcard ga qo'shildi!</i>"
        )

    except ImportError:
        await message.answer("❌ <code>pip install openpyxl</code> kerak!")
    except Exception as e:
        logger.error(f"Universal import error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        await message.answer(f"❌ Xatolik: {e}")
