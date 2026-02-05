"""
Admin Panel Handler - To'liq boshqaruv tizimi
"""
from datetime import datetime, timedelta
from aiogram import Router, F, Bot
from aiogram.types import Message, CallbackQuery
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup

from sqlalchemy import select
from src.database import get_session
from src.database.models import User, Language, Level, Day, Question
from src.repositories import (
    UserRepository, QuestionRepository,
    LanguageRepository, LevelRepository, DayRepository
)
from src.core.logging import get_logger
from src.core.security import is_admin, is_super_admin
from src.handlers.admin.shop_admin import router as shop_admin_router

logger = get_logger(__name__)
router = Router(name="admin")
router.include_router(shop_admin_router)


class AdminStates(StatesGroup):
    """Admin FSM states"""
    waiting_broadcast = State()
    waiting_user_id = State()
    waiting_premium_days = State()
    waiting_search_query = State()
    waiting_block_user_id = State()
    waiting_grant_user_id = State()
    waiting_grant_days = State()
    # Content Management
    waiting_language_name = State()
    waiting_language_flag = State()
    waiting_level_name = State()
    waiting_level_language = State()
    waiting_day_number = State()
    waiting_day_topic = State()
    waiting_day_level = State()
    waiting_question_text = State()
    waiting_question_options = State()
    waiting_question_correct = State()
    waiting_question_explanation = State()
    waiting_delete_question_id = State()
    waiting_edit_question = State()
    # Flashcard Deck Management
    waiting_deck_name = State()
    waiting_deck_description = State()
    waiting_deck_icon = State()
    waiting_deck_price = State()
    waiting_card_front = State()
    waiting_card_back = State()
    waiting_card_example = State()


def admin_menu_keyboard(is_super: bool = False) -> InlineKeyboardMarkup:
    """Admin panel keyboard"""
    builder = InlineKeyboardBuilder()
    
    builder.row(InlineKeyboardButton(text="📊 Statistika", callback_data="admin:stats"))
    builder.row(
        InlineKeyboardButton(text="👥 Foydalanuvchilar", callback_data="admin:users"),
        InlineKeyboardButton(text="📢 Broadcast", callback_data="admin:broadcast")
    )
    builder.row(
        InlineKeyboardButton(text="📚 Tillar", callback_data="admin:languages"),
        InlineKeyboardButton(text="❓ Savollar", callback_data="admin:questions")
    )
    builder.row(
        InlineKeyboardButton(text="💰 To'lovlar", callback_data="admin:payments"),
        InlineKeyboardButton(text="🎁 Promo kodlar", callback_data="admin:promos")
    )
    builder.row(InlineKeyboardButton(text="🛒 Do'kon boshqaruvi", callback_data="admin:shop"))
    
    if is_super:
        builder.row(
            InlineKeyboardButton(text="👑 Admin boshqaruvi", callback_data="admin:manage_admins"),
            InlineKeyboardButton(text="⚙️ Sozlamalar", callback_data="admin:settings")
        )
    
    return builder.as_markup()


def content_menu_keyboard() -> InlineKeyboardMarkup:
    """Content management keyboard"""
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="🌍 Til qo'shish", callback_data="admin:add_language"))
    builder.row(InlineKeyboardButton(text="📊 Daraja qo'shish", callback_data="admin:add_level"))
    builder.row(InlineKeyboardButton(text="📅 Kun qo'shish", callback_data="admin:add_day"))
    builder.row(InlineKeyboardButton(text="❓ Savol qo'shish", callback_data="admin:add_question"))
    builder.row(InlineKeyboardButton(text="📥 Excel'dan import", callback_data="admin:import_excel"))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:panel"))
    return builder.as_markup()


# ============================================================
# ADMIN PANEL ENTRY
# ============================================================

@router.message(Command("admin"))
async def admin_panel_cmd(message: Message):
    """Admin panel command"""
    user_id = message.from_user.id
    
    if not is_admin(user_id):
        await message.answer("❌ Sizda admin huquqi yo'q!")
        return
    
    await show_admin_panel(message, is_super_admin(user_id))


@router.callback_query(F.data == "admin:panel")
async def admin_panel_callback(callback: CallbackQuery):
    """Admin panel callback"""
    user_id = callback.from_user.id
    
    if not is_admin(user_id):
        await callback.answer("❌ Sizda admin huquqi yo'q!", show_alert=True)
        return
    
    await show_admin_panel(callback.message, is_super_admin(user_id), edit=True)
    await callback.answer()


async def show_admin_panel(message: Message, is_super: bool, edit: bool = False):
    """Show admin panel"""
    async with get_session() as session:
        user_repo = UserRepository(session)
        question_repo = QuestionRepository(session)
        
        total_users = await user_repo.count_all()
        premium_users = await user_repo.count_premium()
        total_questions = await question_repo.count()
    
    role = "👑 Super Admin" if is_super else "🔧 Admin"
    
    text = f"""
{role} <b>Panel</b>

📊 <b>Qisqa statistika:</b>
• Foydalanuvchilar: {total_users}
• Premium: {premium_users}
• Savollar: {total_questions}

Boshqarish uchun quyidagi tugmalarni tanlang:
"""
    
    if edit:
        await message.edit_text(text, reply_markup=admin_menu_keyboard(is_super))
    else:
        await message.answer(text, reply_markup=admin_menu_keyboard(is_super))


# ============================================================
# SETTINGS (NEW!)
# ============================================================

@router.callback_query(F.data == "admin:settings")
async def admin_settings(callback: CallbackQuery):
    """Admin settings menu"""
    if not is_super_admin(callback.from_user.id):
        await callback.answer("❌ Faqat Super Admin!", show_alert=True)
        return
    
    text = """
⚙️ <b>Bot Sozlamalari</b>

<b>Joriy sozlamalar:</b>
• Quiz vaqti: {time}s
• Rate limit: {rate}/min
• Streak reset: {streak}h
• Audio: {audio}

<i>Sozlamalarni o'zgartirish uchun .env faylini tahrirlang</i>
""".format(
        time=settings.DEFAULT_TIME_PER_QUESTION,
        rate=settings.RATE_LIMIT_MESSAGES_PER_MINUTE,
        streak=settings.STREAK_RESET_HOURS,
        audio="✅" if settings.AUDIO_ENABLED else "❌"
    )
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="🔄 Cache tozalash", callback_data="admin:clear_cache"))
    builder.row(InlineKeyboardButton(text="📊 System info", callback_data="admin:system_info"))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:panel"))
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.callback_query(F.data == "admin:clear_cache")
async def clear_cache(callback: CallbackQuery):
    """Clear Redis cache"""
    if not is_super_admin(callback.from_user.id):
        return
    
    try:
        from src.core.redis import get_redis
        redis = await get_redis()
        await redis.flushdb()
        await callback.answer("✅ Cache tozalandi!", show_alert=True)
    except Exception as e:
        await callback.answer(f"❌ Xatolik: {e}", show_alert=True)


@router.callback_query(F.data == "admin:system_info")
async def system_info(callback: CallbackQuery):
    """Show system information"""
    if not is_super_admin(callback.from_user.id):
        return
    
    import platform
    import sys
    
    text = f"""
🖥 <b>System Info</b>

• Python: {sys.version.split()[0]}
• OS: {platform.system()} {platform.release()}
• aiogram: 3.x

<b>Database:</b>
• URL: {settings.DATABASE_URL.split('@')[-1] if '@' in settings.DATABASE_URL else 'local'}

<b>Redis:</b>
• URL: {settings.REDIS_URL}
"""
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:settings"))
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


# ============================================================
# STATISTICS
# ============================================================

@router.callback_query(F.data == "admin:stats")
async def admin_stats(callback: CallbackQuery):
    """Detailed statistics"""
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    async with get_session() as session:
        user_repo = UserRepository(session)
        question_repo = QuestionRepository(session)
        lang_repo = LanguageRepository(session)
        
        total_users = await user_repo.count_all()
        premium_users = await user_repo.count_premium()
        today_users = await user_repo.count_today()
        week_users = await user_repo.count_week()
        total_questions = await question_repo.count()
        languages = await lang_repo.get_active_languages()
    
    # Safe division
    if total_users > 0:
        conversion = premium_users / total_users * 100
    else:
        conversion = 0
    
    text = f"""
📊 <b>Batafsil statistika</b>

<b>👥 Foydalanuvchilar:</b>
• Jami: {total_users}
• Premium: {premium_users} ({conversion:.1f}% konversiya)
• Bugun qo'shilgan: {today_users}
• Bu hafta: {week_users}

<b>📚 Kontent:</b>
• Savollar: {total_questions}
• Tillar: {len(languages)}

<b>🌍 Tillar bo'yicha:</b>
"""
    
    for lang in languages:
        text += f"• {lang.flag} {lang.name}\n"
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="📈 Grafik", callback_data="admin:stats_chart"))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:panel"))
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.callback_query(F.data == "admin:stats_chart")
async def stats_chart(callback: CallbackQuery):
    """Statistics chart placeholder"""
    await callback.answer("📈 Grafik tez orada qo'shiladi!", show_alert=True)


# ============================================================
# USER MANAGEMENT
# ============================================================

@router.callback_query(F.data == "admin:users")
async def user_management(callback: CallbackQuery):
    """User management menu"""
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    text = """
👥 <b>Foydalanuvchilar boshqaruvi</b>

Quyidagi amallardan birini tanlang:
"""
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="🔍 Qidirish", callback_data="admin:search_user"))
    builder.row(InlineKeyboardButton(text="👑 Premium berish", callback_data="admin:grant_premium"))
    builder.row(InlineKeyboardButton(text="🚫 Bloklash", callback_data="admin:block_user"))
    builder.row(InlineKeyboardButton(text="📋 Oxirgi ro'yxat", callback_data="admin:recent_users"))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:panel"))
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


# ============== SEARCH USER (NEW!) ==============
@router.callback_query(F.data == "admin:search_user")
async def search_user_start(callback: CallbackQuery, state: FSMContext):
    """Start user search"""
    if not is_admin(callback.from_user.id):
        return
    
    await state.set_state(AdminStates.waiting_search_query)
    
    await callback.message.edit_text(
        "🔍 <b>Foydalanuvchi qidirish</b>\n\n"
        "User ID yoki @username kiriting:\n\n"
        "<i>Masalan: 123456789 yoki @username</i>\n\n"
        "Bekor qilish: /cancel"
    )
    await callback.answer()


@router.message(AdminStates.waiting_search_query)
async def process_search_query(message: Message, state: FSMContext):
    """Process search query"""
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi. /admin")
        return
    
    query = message.text.strip().replace("@", "")
    
    async with get_session() as session:
        user_repo = UserRepository(session)
        
        # Try to find by ID
        if query.isdigit():
            user = await user_repo.get_by_user_id(int(query))
        else:
            # Search by username
            user = await user_repo.get_by_username(query)
    
    await state.clear()
    
    if not user:
        await message.answer(
            "❌ Foydalanuvchi topilmadi!\n\n"
            "Qayta qidirish: /admin → Foydalanuvchilar → Qidirish"
        )
        return
    
    premium_status = "✅ Premium" if user.is_premium else "❌ Oddiy"
    blocked_status = "🚫 BLOKLANGAN" if user.is_blocked else ""
    
    text = f"""
👤 <b>Foydalanuvchi ma'lumotlari</b>

• ID: <code>{user.user_id}</code>
• Ism: {user.full_name}
• Username: @{user.username or "yo'q"}
• Status: {premium_status} {blocked_status}

📊 <b>Statistika:</b>
• Quizlar: {user.total_quizzes}
• To'g'ri javoblar: {user.total_correct}
• Aniqlik: {user.accuracy:.1f}%

📅 Ro'yxatdan o'tgan: {user.created_at.strftime('%Y-%m-%d') if user.created_at else 'N/A'}
"""
    
    builder = InlineKeyboardBuilder()
    if not user.is_premium:
        builder.row(InlineKeyboardButton(
            text="👑 Premium berish",
            callback_data=f"admin:give_premium:{user.user_id}"
        ))
    if user.is_blocked:
        builder.row(InlineKeyboardButton(
            text="✅ Blokdan chiqarish",
            callback_data=f"admin:unblock:{user.user_id}"
        ))
    else:
        builder.row(InlineKeyboardButton(
            text="🚫 Bloklash",
            callback_data=f"admin:do_block:{user.user_id}"
        ))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:users"))
    
    await message.answer(text, reply_markup=builder.as_markup())


# ============== GRANT PREMIUM (NEW!) ==============
@router.callback_query(F.data == "admin:grant_premium")
async def grant_premium_start(callback: CallbackQuery, state: FSMContext):
    """Start granting premium"""
    if not is_super_admin(callback.from_user.id):
        await callback.answer("❌ Faqat Super Admin!", show_alert=True)
        return
    
    await state.set_state(AdminStates.waiting_grant_user_id)
    
    await callback.message.edit_text(
        "👑 <b>Premium berish</b>\n\n"
        "Foydalanuvchi ID sini kiriting:\n\n"
        "Bekor qilish: /cancel"
    )
    await callback.answer()


@router.message(AdminStates.waiting_grant_user_id)
async def receive_grant_user_id(message: Message, state: FSMContext):
    """Receive user ID for premium"""
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi. /admin")
        return
    
    if not message.text.isdigit():
        await message.answer("❌ Faqat raqam kiriting!")
        return
    
    user_id = int(message.text)
    
    async with get_session() as session:
        user_repo = UserRepository(session)
        user = await user_repo.get_by_user_id(user_id)
    
    if not user:
        await message.answer("❌ Foydalanuvchi topilmadi!")
        return
    
    await state.update_data(grant_user_id=user_id, grant_user_name=user.full_name)
    await state.set_state(AdminStates.waiting_grant_days)
    
    await message.answer(
        f"👤 Foydalanuvchi: <b>{user.full_name}</b>\n\n"
        "Necha kun Premium berasiz?\n\n"
        "<i>Masalan: 30</i>"
    )


@router.message(AdminStates.waiting_grant_days)
async def receive_grant_days(message: Message, state: FSMContext):
    """Receive days and grant premium"""
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi. /admin")
        return
    
    if not message.text.isdigit():
        await message.answer("❌ Faqat raqam kiriting!")
        return
    
    days = int(message.text)
    data = await state.get_data()
    user_id = data.get("grant_user_id")
    user_name = data.get("grant_user_name")
    
    async with get_session() as session:
        user_repo = UserRepository(session)
        user = await user_repo.get_by_user_id(user_id)
        
        if user:
            user.is_premium = True
            await session.commit()
    
    await state.clear()
    
    await message.answer(
        f"✅ <b>{user_name}</b> ga {days} kun Premium berildi!\n\n"
        "/admin - Admin panel"
    )


@router.callback_query(F.data.startswith("admin:give_premium:"))
async def give_premium_quick(callback: CallbackQuery, state: FSMContext):
    """Quick premium grant from search result"""
    if not is_super_admin(callback.from_user.id):
        await callback.answer("❌ Faqat Super Admin!", show_alert=True)
        return
    
    user_id = int(callback.data.split(":")[-1])
    await state.update_data(grant_user_id=user_id)
    await state.set_state(AdminStates.waiting_grant_days)
    
    await callback.message.edit_text(
        "👑 <b>Premium berish</b>\n\n"
        "Necha kun Premium berasiz?\n\n"
        "<i>Masalan: 30</i>\n\n"
        "Bekor qilish: /cancel"
    )
    await callback.answer()


# ============== BLOCK USER (NEW!) ==============
@router.callback_query(F.data == "admin:block_user")
async def block_user_start(callback: CallbackQuery, state: FSMContext):
    """Start blocking user"""
    if not is_admin(callback.from_user.id):
        return
    
    await state.set_state(AdminStates.waiting_block_user_id)
    
    await callback.message.edit_text(
        "🚫 <b>Foydalanuvchini bloklash</b>\n\n"
        "Foydalanuvchi ID sini kiriting:\n\n"
        "Bekor qilish: /cancel"
    )
    await callback.answer()


@router.message(AdminStates.waiting_block_user_id)
async def receive_block_user_id(message: Message, state: FSMContext):
    """Receive user ID and block"""
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi. /admin")
        return
    
    if not message.text.isdigit():
        await message.answer("❌ Faqat raqam kiriting!")
        return
    
    user_id = int(message.text)
    
    async with get_session() as session:
        user_repo = UserRepository(session)
        user = await user_repo.get_by_user_id(user_id)
        
        if not user:
            await message.answer("❌ Foydalanuvchi topilmadi!")
            return
        
        user.is_blocked = True
        await session.commit()
    
    await state.clear()
    await message.answer(f"🚫 <b>{user.full_name}</b> bloklandi!\n\n/admin")


@router.callback_query(F.data.startswith("admin:do_block:"))
async def do_block_quick(callback: CallbackQuery):
    """Quick block from search result"""
    if not is_admin(callback.from_user.id):
        return
    
    user_id = int(callback.data.split(":")[-1])
    
    async with get_session() as session:
        user_repo = UserRepository(session)
        user = await user_repo.get_by_user_id(user_id)
        
        if user:
            user.is_blocked = True
            await session.commit()
            await callback.answer(f"🚫 {user.full_name} bloklandi!", show_alert=True)
        else:
            await callback.answer("❌ Foydalanuvchi topilmadi!", show_alert=True)


@router.callback_query(F.data.startswith("admin:unblock:"))
async def unblock_user(callback: CallbackQuery):
    """Unblock user"""
    if not is_admin(callback.from_user.id):
        return
    
    user_id = int(callback.data.split(":")[-1])
    
    async with get_session() as session:
        user_repo = UserRepository(session)
        user = await user_repo.get_by_user_id(user_id)
        
        if user:
            user.is_blocked = False
            await session.commit()
            await callback.answer(f"✅ {user.full_name} blokdan chiqarildi!", show_alert=True)


@router.callback_query(F.data == "admin:recent_users")
async def recent_users(callback: CallbackQuery):
    """Show recent users"""
    if not is_admin(callback.from_user.id):
        return
    
    async with get_session() as session:
        user_repo = UserRepository(session)
        users = await user_repo.get_recent(limit=15)
    
    text = "👥 <b>Oxirgi foydalanuvchilar</b>\n\n"
    
    for user in users:
        premium = "⭐" if user.is_premium else ""
        text += f"• {user.full_name} (@{user.username or 'N/A'}) {premium}\n"
        text += f"  ID: <code>{user.user_id}</code>\n"
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:users"))
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


# Command handlers for backward compatibility
@router.message(Command("grant"))
async def grant_premium_cmd(message: Message):
    """Grant premium command"""
    if not is_super_admin(message.from_user.id):
        await message.answer("❌ Faqat Super Admin!")
        return
    
    args = message.text.split()[1:]
    
    if len(args) < 2:
        await message.answer("❌ Format: /grant [user_id] [days]\nMasalan: /grant 123456789 30")
        return
    
    try:
        user_id = int(args[0])
        days = int(args[1])
    except ValueError:
        await message.answer("❌ Noto'g'ri format!")
        return
    
    async with get_session() as session:
        user_repo = UserRepository(session)
        user = await user_repo.get_by_user_id(user_id)
        
        if not user:
            await message.answer("❌ Foydalanuvchi topilmadi!")
            return
        
        user.is_premium = True
        await session.commit()
    
    await message.answer(f"✅ {user.full_name} ga {days} kun Premium berildi!")


@router.message(Command("block"))
async def block_user_cmd(message: Message):
    """Block user command"""
    if not is_admin(message.from_user.id):
        return
    
    args = message.text.split()[1:]
    
    if not args:
        await message.answer("❌ Format: /block [user_id]")
        return
    
    try:
        user_id = int(args[0])
    except ValueError:
        await message.answer("❌ Noto'g'ri user_id!")
        return
    
    async with get_session() as session:
        user_repo = UserRepository(session)
        user = await user_repo.get_by_user_id(user_id)
        
        if not user:
            await message.answer("❌ Foydalanuvchi topilmadi!")
            return
        
        user.is_blocked = True
        await session.commit()
    
    await message.answer(f"🚫 {user.full_name} bloklandi!")


# ============================================================
# CONTENT MANAGEMENT (QUESTIONS)
# ============================================================

@router.callback_query(F.data == "admin:questions")
async def questions_menu(callback: CallbackQuery):
    """Questions management menu"""
    if not is_admin(callback.from_user.id):
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    async with get_session() as session:
        question_repo = QuestionRepository(session)
        lang_repo = LanguageRepository(session)
        
        total = await question_repo.count()
        languages = await lang_repo.get_active_languages()
    
    text = f"""
❓ <b>Savollar boshqaruvi</b>

📊 Jami savollar: {total}

<b>Tillar bo'yicha:</b>
"""
    
    for lang in languages:
        text += f"• {lang.flag} {lang.name}\n"
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="➕ Savol qo'shish", callback_data="admin:add_question"))
    builder.row(InlineKeyboardButton(text="📥 Excel import", callback_data="admin:import_excel"))
    builder.row(InlineKeyboardButton(text="📋 Savollar ro'yxati", callback_data="admin:list_questions"))
    builder.row(InlineKeyboardButton(text="🗑 Savol o'chirish", callback_data="admin:delete_question"))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:panel"))
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


# ============== LIST QUESTIONS (NEW!) ==============
@router.callback_query(F.data == "admin:list_questions")
async def list_questions(callback: CallbackQuery):
    """List questions - show topics (Days) directly"""
    if not is_admin(callback.from_user.id):
        return

    async with get_session() as session:
        # Get all Days with question counts
        from src.database.models import Day, Level
        result = await session.execute(
            select(Day, Level)
            .join(Level, Day.level_id == Level.id)
            .where(Day.is_active == True)
            .order_by(Level.display_order, Day.day_number)
        )
        days_with_levels = result.all()

    if not days_with_levels:
        await callback.answer("❌ Mavzular yo'q!", show_alert=True)
        return

    # Group by level
    by_level = {}
    for day, level in days_with_levels:
        if level.name not in by_level:
            by_level[level.name] = []
        by_level[level.name].append(day)

    text = "📋 <b>Savollar ro'yxati</b>\n\nMavzuni tanlang:\n"

    builder = InlineKeyboardBuilder()
    for level_name, days in by_level.items():
        # Add level header as a row
        for day in days[:10]:  # First 10 days per level
            builder.row(InlineKeyboardButton(
                text=f"{level_name} | {day.name[:25]}",
                callback_data=f"admin:list_q_day:{day.id}"
            ))

    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:questions"))

    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.callback_query(F.data.startswith("admin:list_q_day:"))
async def list_questions_by_day(callback: CallbackQuery):
    """List questions by day/topic with pagination"""
    parts = callback.data.split(":")
    day_id = int(parts[2])
    page = int(parts[3]) if len(parts) > 3 else 1

    per_page = 5  # Reduced to show buttons
    offset = (page - 1) * per_page

    async with get_session() as session:
        question_repo = QuestionRepository(session)
        from src.database.models import Day

        # Get day info
        day_result = await session.execute(select(Day).where(Day.id == day_id))
        day = day_result.scalar_one_or_none()

        # Get questions for this day
        all_questions = await question_repo.get_by_day(day_id, active_only=False)
        total = len(all_questions)
        questions = all_questions[offset:offset + per_page]

    if not day:
        await callback.answer("❌ Mavzu topilmadi!", show_alert=True)
        return

    total_pages = max(1, (total + per_page - 1) // per_page)

    text = f"📋 <b>{day.name}</b>\n"
    text += f"📊 Jami: {total} ta savol | Sahifa {page}/{total_pages}\n\n"
    text += "<i>Savolni bosib tahrirlang:</i>\n"

    builder = InlineKeyboardBuilder()

    if questions:
        for i, q in enumerate(questions, offset + 1):
            short_text = q.question_text[:30] + "..." if len(q.question_text) > 30 else q.question_text
            status = "✅" if q.is_active else "❌"
            builder.row(InlineKeyboardButton(
                text=f"{i}. {status} {short_text}",
                callback_data=f"admin:view_q:{q.id}:{day_id}:{page}"
            ))
    else:
        text += "<i>Savollar yo'q</i>"

    # Pagination buttons
    nav_buttons = []
    if page > 1:
        nav_buttons.append(InlineKeyboardButton(
            text="◀️ Oldingi",
            callback_data=f"admin:list_q_day:{day_id}:{page-1}"
        ))
    if page < total_pages:
        nav_buttons.append(InlineKeyboardButton(
            text="Keyingi ▶️",
            callback_data=f"admin:list_q_day:{day_id}:{page+1}"
        ))

    if nav_buttons:
        builder.row(*nav_buttons)

    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:list_questions"))

    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


# ============== VIEW & EDIT QUESTION ==============
@router.callback_query(F.data.startswith("admin:view_q:"))
async def view_question(callback: CallbackQuery):
    """View full question details"""
    parts = callback.data.split(":")
    question_id = int(parts[2])
    day_id = int(parts[3]) if len(parts) > 3 else 0
    page = int(parts[4]) if len(parts) > 4 else 1

    async with get_session() as session:
        question_repo = QuestionRepository(session)
        question = await question_repo.get_by_id(question_id)

    if not question:
        await callback.answer("❌ Savol topilmadi!", show_alert=True)
        return

    status = "✅ Faol" if question.is_active else "❌ Nofaol"

    text = f"""❓ <b>Savol #{question.id}</b>

<b>Savol matni:</b>
{question.question_text}

<b>Variantlar:</b>
A) {question.option_a}
B) {question.option_b}
C) {question.option_c}
D) {question.option_d}

<b>To'g'ri javob:</b> {question.correct_option}

<b>Tushuntirish:</b>
{question.explanation or "—"}

<b>Statistika:</b>
👁 Ko'rsatildi: {question.times_shown} marta
✅ To'g'ri javoblar: {question.times_correct}
👍 Upvotes: {question.upvotes} | 👎 Downvotes: {question.downvotes}

<b>Holat:</b> {status}
"""

    builder = InlineKeyboardBuilder()

    # Edit buttons
    builder.row(
        InlineKeyboardButton(text="✏️ Savol", callback_data=f"admin:edit_q:text:{question_id}"),
        InlineKeyboardButton(text="✏️ Tushuntirish", callback_data=f"admin:edit_q:expl:{question_id}")
    )
    builder.row(
        InlineKeyboardButton(text="✏️ A", callback_data=f"admin:edit_q:a:{question_id}"),
        InlineKeyboardButton(text="✏️ B", callback_data=f"admin:edit_q:b:{question_id}"),
        InlineKeyboardButton(text="✏️ C", callback_data=f"admin:edit_q:c:{question_id}"),
        InlineKeyboardButton(text="✏️ D", callback_data=f"admin:edit_q:d:{question_id}")
    )
    builder.row(
        InlineKeyboardButton(text="✏️ To'g'ri javob", callback_data=f"admin:edit_q:correct:{question_id}")
    )

    # Toggle active
    if question.is_active:
        builder.row(InlineKeyboardButton(
            text="❌ Nofaol qilish",
            callback_data=f"admin:toggle_q:{question_id}:{day_id}:{page}"
        ))
    else:
        builder.row(InlineKeyboardButton(
            text="✅ Faol qilish",
            callback_data=f"admin:toggle_q:{question_id}:{day_id}:{page}"
        ))

    # Delete button
    builder.row(InlineKeyboardButton(
        text="🗑 O'chirish",
        callback_data=f"admin:del_q:{question_id}:{day_id}:{page}"
    ))

    # Back button
    builder.row(InlineKeyboardButton(
        text="◀️ Orqaga",
        callback_data=f"admin:list_q_day:{day_id}:{page}"
    ))

    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.callback_query(F.data.startswith("admin:toggle_q:"))
async def toggle_question(callback: CallbackQuery):
    """Toggle question active status"""
    parts = callback.data.split(":")
    question_id = int(parts[2])
    day_id = int(parts[3]) if len(parts) > 3 else 0
    page = int(parts[4]) if len(parts) > 4 else 1

    async with get_session() as session:
        question_repo = QuestionRepository(session)
        question = await question_repo.get_by_id(question_id)
        if question:
            question.is_active = not question.is_active
            await session.commit()
            status = "faollashtirildi" if question.is_active else "nofaol qilindi"
            await callback.answer(f"✅ Savol {status}!", show_alert=True)

    # Refresh view
    callback.data = f"admin:view_q:{question_id}:{day_id}:{page}"
    await view_question(callback)


@router.callback_query(F.data.startswith("admin:del_q:"))
async def delete_question_confirm(callback: CallbackQuery):
    """Confirm question deletion"""
    parts = callback.data.split(":")
    question_id = int(parts[2])
    day_id = int(parts[3]) if len(parts) > 3 else 0
    page = int(parts[4]) if len(parts) > 4 else 1

    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(
            text="✅ Ha, o'chirish",
            callback_data=f"admin:del_q_yes:{question_id}:{day_id}:{page}"
        ),
        InlineKeyboardButton(
            text="❌ Yo'q",
            callback_data=f"admin:view_q:{question_id}:{day_id}:{page}"
        )
    )

    await callback.message.edit_text(
        "⚠️ <b>Savolni o'chirishni tasdiqlaysizmi?</b>\n\n"
        "Bu amalni qaytarib bo'lmaydi!",
        reply_markup=builder.as_markup()
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin:del_q_yes:"))
async def delete_question_execute(callback: CallbackQuery):
    """Execute question deletion"""
    parts = callback.data.split(":")
    question_id = int(parts[2])
    day_id = int(parts[3]) if len(parts) > 3 else 0
    page = int(parts[4]) if len(parts) > 4 else 1

    async with get_session() as session:
        question_repo = QuestionRepository(session)
        question = await question_repo.get_by_id(question_id)
        if question:
            await session.delete(question)
            await session.commit()

    await callback.answer("🗑 Savol o'chirildi!", show_alert=True)

    # Go back to list
    callback.data = f"admin:list_q_day:{day_id}:{page}"
    await list_questions_by_day(callback)


@router.callback_query(F.data.startswith("admin:edit_q:"))
async def edit_question_start(callback: CallbackQuery, state: FSMContext):
    """Start editing a question field"""
    parts = callback.data.split(":")
    field = parts[2]  # text, expl, a, b, c, d, correct
    question_id = int(parts[3])

    field_names = {
        "text": "Savol matni",
        "expl": "Tushuntirish",
        "a": "A varianti",
        "b": "B varianti",
        "c": "C varianti",
        "d": "D varianti",
        "correct": "To'g'ri javob (A, B, C yoki D)"
    }

    await state.set_state(AdminStates.waiting_edit_question)
    await state.update_data(edit_q_id=question_id, edit_q_field=field)

    text = f"✏️ <b>{field_names.get(field, field)}</b> ni kiriting:\n\n"
    if field == "correct":
        text += "<i>Faqat A, B, C yoki D kiriting</i>\n\n"
    text += "Bekor qilish: /cancel"

    await callback.message.edit_text(text)
    await callback.answer()


@router.message(AdminStates.waiting_edit_question)
async def process_edit_question(message: Message, state: FSMContext):
    """Process question edit"""
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi. /admin")
        return

    data = await state.get_data()
    question_id = data.get("edit_q_id")
    field = data.get("edit_q_field")
    new_value = message.text.strip()

    # Validate correct answer
    if field == "correct":
        new_value = new_value.upper()
        if new_value not in ["A", "B", "C", "D"]:
            await message.answer("❌ Faqat A, B, C yoki D kiriting!")
            return

    async with get_session() as session:
        question_repo = QuestionRepository(session)
        question = await question_repo.get_by_id(question_id)

        if question:
            if field == "text":
                question.question_text = new_value
            elif field == "expl":
                question.explanation = new_value
            elif field == "a":
                question.option_a = new_value
            elif field == "b":
                question.option_b = new_value
            elif field == "c":
                question.option_c = new_value
            elif field == "d":
                question.option_d = new_value
            elif field == "correct":
                question.correct_option = new_value

            await session.commit()
            day_id = question.day_id

    await state.clear()

    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(
        text="👁 Savolni ko'rish",
        callback_data=f"admin:view_q:{question_id}:{day_id}:1"
    ))
    builder.row(InlineKeyboardButton(
        text="◀️ Admin panel",
        callback_data="admin:questions"
    ))

    await message.answer(
        "✅ <b>Saqlandi!</b>",
        reply_markup=builder.as_markup()
    )


# ============== DELETE QUESTION (NEW!) ==============
@router.callback_query(F.data == "admin:delete_question")
async def delete_question_start(callback: CallbackQuery, state: FSMContext):
    """Start deleting question"""
    if not is_admin(callback.from_user.id):
        return
    
    await state.set_state(AdminStates.waiting_delete_question_id)
    
    await callback.message.edit_text(
        "🗑 <b>Savol o'chirish</b>\n\n"
        "Savol ID sini kiriting:\n\n"
        "<i>ID ni \"Savollar ro'yxati\"dan olishingiz mumkin</i>\n\n"
        "Bekor qilish: /cancel"
    )
    await callback.answer()


@router.message(AdminStates.waiting_delete_question_id)
async def process_delete_question(message: Message, state: FSMContext):
    """Process question deletion"""
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi. /admin")
        return
    
    if not message.text.isdigit():
        await message.answer("❌ Faqat raqam kiriting!")
        return
    
    question_id = int(message.text)
    
    async with get_session() as session:
        question_repo = QuestionRepository(session)
        question = await question_repo.get_by_id(question_id)
        
        if not question:
            await message.answer("❌ Savol topilmadi!")
            return
        
        await question_repo.delete(question_id)
        await session.commit()
    
    await state.clear()
    await message.answer(
        f"✅ Savol o'chirildi!\n\n"
        f"ID: {question_id}\n\n"
        "/admin - Admin panel"
    )


@router.callback_query(F.data == "admin:add_question")
async def start_add_question(callback: CallbackQuery, state: FSMContext):
    """Start adding question"""
    if not is_admin(callback.from_user.id):
        return
    
    # Get languages for selection
    async with get_session() as session:
        lang_repo = LanguageRepository(session)
        languages = await lang_repo.get_active_languages()
    
    if not languages:
        await callback.answer("❌ Avval til qo'shing!", show_alert=True)
        return
    
    builder = InlineKeyboardBuilder()
    for lang in languages:
        builder.row(InlineKeyboardButton(
            text=f"{lang.flag} {lang.name}",
            callback_data=f"admin:q_lang:{lang.id}"
        ))
    builder.row(InlineKeyboardButton(text="❌ Bekor", callback_data="admin:questions"))
    
    await callback.message.edit_text(
        "❓ <b>Savol qo'shish</b>\n\n"
        "1️⃣ Tilni tanlang:",
        reply_markup=builder.as_markup()
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin:q_lang:"))
async def select_question_language(callback: CallbackQuery, state: FSMContext):
    """Select language for question"""
    lang_id = int(callback.data.split(":")[-1])
    await state.update_data(question_lang_id=lang_id)
    
    # Get levels
    async with get_session() as session:
        level_repo = LevelRepository(session)
        levels = await level_repo.get_by_language(lang_id)
    
    if not levels:
        await callback.answer("❌ Bu til uchun daraja yo'q! Avval daraja qo'shing.", show_alert=True)
        return
    
    builder = InlineKeyboardBuilder()
    for level in levels:
        builder.row(InlineKeyboardButton(
            text=f"📊 {level.name}",
            callback_data=f"admin:q_level:{level.id}"
        ))
    builder.row(InlineKeyboardButton(text="❌ Bekor", callback_data="admin:questions"))
    
    await callback.message.edit_text(
        "❓ <b>Savol qo'shish</b>\n\n"
        "2️⃣ Darajani tanlang:",
        reply_markup=builder.as_markup()
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin:q_level:"))
async def select_question_level(callback: CallbackQuery, state: FSMContext):
    """Select level for question"""
    level_id = int(callback.data.split(":")[-1])
    await state.update_data(question_level_id=level_id)
    
    # Get days
    async with get_session() as session:
        day_repo = DayRepository(session)
        days = await day_repo.get_by_level(level_id)
    
    if not days:
        await callback.answer("❌ Bu daraja uchun kun yo'q! Avval kun qo'shing.", show_alert=True)
        return
    
    builder = InlineKeyboardBuilder()
    for day in days:
        builder.row(InlineKeyboardButton(
            text=f"📅 Kun {day.day_number}: {day.topic}",
            callback_data=f"admin:q_day:{day.id}"
        ))
    builder.row(InlineKeyboardButton(text="❌ Bekor", callback_data="admin:questions"))
    
    await callback.message.edit_text(
        "❓ <b>Savol qo'shish</b>\n\n"
        "3️⃣ Kunni tanlang:",
        reply_markup=builder.as_markup()
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin:q_day:"))
async def select_question_day(callback: CallbackQuery, state: FSMContext):
    """Select day and start question input"""
    day_id = int(callback.data.split(":")[-1])
    await state.update_data(question_day_id=day_id)
    await state.set_state(AdminStates.waiting_question_text)
    
    await callback.message.edit_text(
        "❓ <b>Savol qo'shish</b>\n\n"
        "4️⃣ Savol matnini yozing:\n\n"
        "<i>Masalan: \"Guten Morgen\" qanday tarjima qilinadi?</i>\n\n"
        "Bekor qilish: /cancel"
    )
    await callback.answer()


@router.message(AdminStates.waiting_question_text)
async def receive_question_text(message: Message, state: FSMContext):
    """Receive question text"""
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi. /admin")
        return
    
    await state.update_data(question_text=message.text)
    await state.set_state(AdminStates.waiting_question_options)
    
    await message.answer(
        "❓ <b>Savol qo'shish</b>\n\n"
        "5️⃣ Variantlarni yozing (har birini yangi qatordan):\n\n"
        "<i>Masalan:\n"
        "Xayrli tong\n"
        "Xayrli kech\n"
        "Salom\n"
        "Xayr</i>"
    )


@router.message(AdminStates.waiting_question_options)
async def receive_question_options(message: Message, state: FSMContext):
    """Receive question options"""
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi. /admin")
        return
    
    options = [opt.strip() for opt in message.text.split("\n") if opt.strip()]
    
    if len(options) < 2:
        await message.answer("❌ Kamida 2 ta variant bo'lishi kerak!")
        return
    
    if len(options) > 6:
        await message.answer("❌ Maksimum 6 ta variant!")
        return
    
    await state.update_data(question_options=options)
    await state.set_state(AdminStates.waiting_question_correct)
    
    text = "❓ <b>Savol qo'shish</b>\n\n6️⃣ To'g'ri javob raqamini yozing:\n\n"
    for i, opt in enumerate(options, 1):
        text += f"{i}. {opt}\n"
    
    await message.answer(text)


@router.message(AdminStates.waiting_question_correct)
async def receive_question_correct(message: Message, state: FSMContext):
    """Receive correct answer index"""
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi. /admin")
        return
    
    data = await state.get_data()
    options = data.get("question_options", [])
    
    try:
        correct_idx = int(message.text) - 1
        if correct_idx < 0 or correct_idx >= len(options):
            raise ValueError()
    except ValueError:
        await message.answer(f"❌ 1 dan {len(options)} gacha raqam kiriting!")
        return
    
    await state.update_data(question_correct=correct_idx)
    await state.set_state(AdminStates.waiting_question_explanation)
    
    await message.answer(
        "❓ <b>Savol qo'shish</b>\n\n"
        "7️⃣ Tushuntirish yozing (ixtiyoriy):\n\n"
        "<i>Masalan: \"Guten Morgen\" - nemischa \"Xayrli tong\" degani.</i>\n\n"
        "O'tkazib yuborish uchun: /skip"
    )


@router.message(AdminStates.waiting_question_explanation)
async def receive_question_explanation(message: Message, state: FSMContext):
    """Receive explanation and save question"""
    data = await state.get_data()
    
    explanation = None if message.text == "/skip" else message.text
    
    # Save question
    async with get_session() as session:
        question = Question(
            day_id=data["question_day_id"],
            question_text=data["question_text"],
            correct_answer=data["question_options"][data["question_correct"]],
            wrong_answers=data["question_options"],  # Will be filtered
            explanation=explanation,
            is_active=True
        )
        
        # Set correct wrong_answers (excluding correct one)
        all_options = data["question_options"]
        correct_idx = data["question_correct"]
        question.wrong_answers = [opt for i, opt in enumerate(all_options) if i != correct_idx]
        
        session.add(question)
        await session.commit()
    
    await state.clear()
    
    await message.answer(
        "✅ <b>Savol qo'shildi!</b>\n\n"
        f"📝 {data['question_text']}\n"
        f"✓ To'g'ri: {data['question_options'][data['question_correct']]}\n\n"
        "Yana savol qo'shish: /admin → Savollar → Savol qo'shish"
    )


# ============================================================
# LANGUAGE MANAGEMENT
# ============================================================

@router.callback_query(F.data == "admin:languages")
async def languages_menu(callback: CallbackQuery):
    """Languages management"""
    if not is_admin(callback.from_user.id):
        return
    
    async with get_session() as session:
        lang_repo = LanguageRepository(session)
        languages = await lang_repo.get_active_languages()
    
    text = "🌍 <b>Tillar boshqaruvi</b>\n\n"
    
    if languages:
        for lang in languages:
            text += f"• {lang.flag} {lang.name} (ID: {lang.id})\n"
    else:
        text += "<i>Tillar yo'q</i>\n"
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="➕ Til qo'shish", callback_data="admin:add_language"))
    builder.row(InlineKeyboardButton(text="📊 Daraja qo'shish", callback_data="admin:add_level"))
    builder.row(InlineKeyboardButton(text="📅 Kun qo'shish", callback_data="admin:add_day"))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:panel"))
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.callback_query(F.data == "admin:add_language")
async def start_add_language(callback: CallbackQuery, state: FSMContext):
    """Start adding language"""
    if not is_super_admin(callback.from_user.id):
        await callback.answer("❌ Faqat Super Admin!", show_alert=True)
        return
    
    await state.set_state(AdminStates.waiting_language_name)
    
    await callback.message.edit_text(
        "🌍 <b>Til qo'shish</b>\n\n"
        "Til nomini yozing:\n"
        "<i>Masalan: Nemis tili</i>\n\n"
        "Bekor qilish: /cancel"
    )
    await callback.answer()


@router.message(AdminStates.waiting_language_name)
async def receive_language_name(message: Message, state: FSMContext):
    """Receive language name"""
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi. /admin")
        return
    
    await state.update_data(lang_name=message.text)
    await state.set_state(AdminStates.waiting_language_flag)
    
    await message.answer(
        "🌍 <b>Til qo'shish</b>\n\n"
        "Bayroq emoji yuboring:\n"
        "<i>Masalan: 🇩🇪</i>"
    )


@router.message(AdminStates.waiting_language_flag)
async def receive_language_flag(message: Message, state: FSMContext):
    """Receive language flag and save"""
    data = await state.get_data()
    
    async with get_session() as session:
        lang = Language(
            name=data["lang_name"],
            code=data["lang_name"][:2].lower(),
            flag=message.text,
            is_active=True
        )
        session.add(lang)
        await session.commit()
    
    await state.clear()
    await message.answer(
        f"✅ Til qo'shildi: {message.text} {data['lang_name']}\n\n"
        "Endi daraja qo'shish: /admin → Tillar → Daraja qo'shish"
    )


# ============== ADD LEVEL (NEW!) ==============
@router.callback_query(F.data == "admin:add_level")
async def start_add_level(callback: CallbackQuery, state: FSMContext):
    """Start adding level"""
    if not is_admin(callback.from_user.id):
        return
    
    async with get_session() as session:
        lang_repo = LanguageRepository(session)
        languages = await lang_repo.get_active_languages()
    
    if not languages:
        await callback.answer("❌ Avval til qo'shing!", show_alert=True)
        return
    
    builder = InlineKeyboardBuilder()
    for lang in languages:
        builder.row(InlineKeyboardButton(
            text=f"{lang.flag} {lang.name}",
            callback_data=f"admin:level_lang:{lang.id}"
        ))
    builder.row(InlineKeyboardButton(text="❌ Bekor", callback_data="admin:languages"))
    
    await callback.message.edit_text(
        "📊 <b>Daraja qo'shish</b>\n\n"
        "1️⃣ Tilni tanlang:",
        reply_markup=builder.as_markup()
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin:level_lang:"))
async def select_level_language(callback: CallbackQuery, state: FSMContext):
    """Select language for new level"""
    lang_id = int(callback.data.split(":")[-1])
    await state.update_data(level_lang_id=lang_id)
    await state.set_state(AdminStates.waiting_level_name)
    
    await callback.message.edit_text(
        "📊 <b>Daraja qo'shish</b>\n\n"
        "2️⃣ Daraja nomini yozing:\n\n"
        "<i>Masalan: A1, A2, B1, B2 yoki Boshlang'ich</i>\n\n"
        "Bekor qilish: /cancel"
    )
    await callback.answer()


@router.message(AdminStates.waiting_level_name)
async def receive_level_name(message: Message, state: FSMContext):
    """Receive level name and save"""
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi. /admin")
        return
    
    data = await state.get_data()
    lang_id = data.get("level_lang_id")
    
    async with get_session() as session:
        # Get current max order
        level_repo = LevelRepository(session)
        levels = await level_repo.get_by_language(lang_id)
        max_order = max([l.order for l in levels], default=0) + 1
        
        level = Level(
            language_id=lang_id,
            name=message.text,
            order=max_order,
            is_active=True
        )
        session.add(level)
        await session.commit()
    
    await state.clear()
    await message.answer(
        f"✅ Daraja qo'shildi: {message.text}\n\n"
        "Endi kun qo'shish: /admin → Tillar → Kun qo'shish"
    )


# ============== ADD DAY (NEW!) ==============
@router.callback_query(F.data == "admin:add_day")
async def start_add_day(callback: CallbackQuery, state: FSMContext):
    """Start adding day"""
    if not is_admin(callback.from_user.id):
        return
    
    async with get_session() as session:
        lang_repo = LanguageRepository(session)
        languages = await lang_repo.get_active_languages()
    
    if not languages:
        await callback.answer("❌ Avval til qo'shing!", show_alert=True)
        return
    
    builder = InlineKeyboardBuilder()
    for lang in languages:
        builder.row(InlineKeyboardButton(
            text=f"{lang.flag} {lang.name}",
            callback_data=f"admin:day_lang:{lang.id}"
        ))
    builder.row(InlineKeyboardButton(text="❌ Bekor", callback_data="admin:languages"))
    
    await callback.message.edit_text(
        "📅 <b>Kun qo'shish</b>\n\n"
        "1️⃣ Tilni tanlang:",
        reply_markup=builder.as_markup()
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin:day_lang:"))
async def select_day_language(callback: CallbackQuery, state: FSMContext):
    """Select language for day"""
    lang_id = int(callback.data.split(":")[-1])
    await state.update_data(day_lang_id=lang_id)
    
    async with get_session() as session:
        level_repo = LevelRepository(session)
        levels = await level_repo.get_by_language(lang_id)
    
    if not levels:
        await callback.answer("❌ Bu til uchun daraja yo'q! Avval daraja qo'shing.", show_alert=True)
        return
    
    builder = InlineKeyboardBuilder()
    for level in levels:
        builder.row(InlineKeyboardButton(
            text=f"📊 {level.name}",
            callback_data=f"admin:day_level:{level.id}"
        ))
    builder.row(InlineKeyboardButton(text="❌ Bekor", callback_data="admin:languages"))
    
    await callback.message.edit_text(
        "📅 <b>Kun qo'shish</b>\n\n"
        "2️⃣ Darajani tanlang:",
        reply_markup=builder.as_markup()
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin:day_level:"))
async def select_day_level(callback: CallbackQuery, state: FSMContext):
    """Select level for day"""
    level_id = int(callback.data.split(":")[-1])
    await state.update_data(day_level_id=level_id)
    await state.set_state(AdminStates.waiting_day_number)
    
    await callback.message.edit_text(
        "📅 <b>Kun qo'shish</b>\n\n"
        "3️⃣ Kun raqamini yozing:\n\n"
        "<i>Masalan: 1, 2, 3...</i>\n\n"
        "Bekor qilish: /cancel"
    )
    await callback.answer()


@router.message(AdminStates.waiting_day_number)
async def receive_day_number(message: Message, state: FSMContext):
    """Receive day number"""
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi. /admin")
        return
    
    if not message.text.isdigit():
        await message.answer("❌ Faqat raqam kiriting!")
        return
    
    await state.update_data(day_number=int(message.text))
    await state.set_state(AdminStates.waiting_day_topic)
    
    await message.answer(
        "📅 <b>Kun qo'shish</b>\n\n"
        "4️⃣ Kun mavzusini yozing:\n\n"
        "<i>Masalan: Salomlashish, Tanishuv, Oila...</i>"
    )


@router.message(AdminStates.waiting_day_topic)
async def receive_day_topic(message: Message, state: FSMContext):
    """Receive day topic and save"""
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi. /admin")
        return
    
    data = await state.get_data()
    
    async with get_session() as session:
        day = Day(
            level_id=data["day_level_id"],
            day_number=data["day_number"],
            topic=message.text,
            is_active=True
        )
        session.add(day)
        await session.commit()
    
    await state.clear()
    await message.answer(
        f"✅ Kun qo'shildi!\n\n"
        f"📅 Kun {data['day_number']}: {message.text}\n\n"
        "Endi savol qo'shish: /admin → Savollar → Savol qo'shish"
    )


# ============================================================
# BROADCAST
# ============================================================

@router.callback_query(F.data == "admin:broadcast")
async def broadcast_menu(callback: CallbackQuery, state: FSMContext):
    """Start broadcast"""
    if not is_admin(callback.from_user.id):
        return
    
    await state.set_state(AdminStates.waiting_broadcast)
    
    await callback.message.edit_text(
        "📢 <b>Broadcast</b>\n\n"
        "Barcha foydalanuvchilarga yuboriladigan xabarni yozing.\n\n"
        "Rasm yuborish uchun rasm + caption yuboring.\n\n"
        "Bekor qilish: /cancel"
    )
    await callback.answer()


@router.message(AdminStates.waiting_broadcast)
async def process_broadcast(message: Message, state: FSMContext, bot: Bot):
    """Process broadcast"""
    if message.text == "/cancel":
        await state.clear()
        await message.answer("❌ Bekor qilindi")
        return
    
    await state.clear()
    
    broadcast_text = message.text or message.caption or ""
    
    status_msg = await message.answer("📤 Yuborilmoqda... 0%")
    
    async with get_session() as session:
        user_repo = UserRepository(session)
        users = await user_repo.get_all_active(limit=50000)
    
    total = len(users)
    success = 0
    failed = 0
    
    for i, user in enumerate(users):
        try:
            if message.photo:
                await bot.send_photo(user.user_id, message.photo[-1].file_id, caption=broadcast_text)
            else:
                await bot.send_message(user.user_id, broadcast_text)
            success += 1
        except Exception:
            failed += 1

        if i % 100 == 0 and total > 0:
            progress = int((i / total) * 100)
            try:
                await status_msg.edit_text(f"📤 Yuborilmoqda... {progress}%")
            except Exception:
                pass
    
    await status_msg.edit_text(
        f"✅ <b>Broadcast tugadi!</b>\n\n"
        f"📤 Yuborildi: {success}\n"
        f"❌ Xatolik: {failed}"
    )


# ============================================================
# PAYMENTS & PROMOS
# ============================================================

@router.callback_query(F.data == "admin:payments")
async def payments_menu(callback: CallbackQuery):
    """Payments overview"""
    if not is_super_admin(callback.from_user.id):
        await callback.answer("❌ Faqat Super Admin!", show_alert=True)
        return
    
    text = """
💰 <b>To'lovlar</b>

📊 Statistika tez orada qo'shiladi.

<b>Buyruqlar:</b>
• /grant [id] [days] - Premium berish
• /refund [payment_id] - Qaytarish
"""
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:panel"))
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.callback_query(F.data == "admin:promos")
async def promos_menu(callback: CallbackQuery):
    """Promo codes menu"""
    if not is_admin(callback.from_user.id):
        return
    
    text = """
🎁 <b>Promo kodlar</b>

<b>Buyruqlar:</b>
• /addpromo [code] [days] [limit] - Yangi kod
• /delpromo [code] - O'chirish
• /promos - Ro'yxat

<i>Masalan: /addpromo SALE2024 7 100</i>
"""
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:panel"))
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


# ============================================================
# ADMIN MANAGEMENT (SUPER ADMIN ONLY)
# ============================================================

@router.callback_query(F.data == "admin:manage_admins")
async def manage_admins(callback: CallbackQuery):
    """Manage admins (super admin only)"""
    if not is_super_admin(callback.from_user.id):
        await callback.answer("❌ Faqat Super Admin!", show_alert=True)
        return
    
    admin_ids = settings.ADMIN_IDS
    super_ids = settings.SUPER_ADMIN_IDS
    
    text = "👑 <b>Admin boshqaruvi</b>\n\n"
    text += "<b>Super Adminlar:</b>\n"
    for uid in super_ids:
        text += f"• <code>{uid}</code>\n"
    
    text += "\n<b>Adminlar:</b>\n"
    for uid in admin_ids:
        if uid not in super_ids:
            text += f"• <code>{uid}</code>\n"
    
    text += "\n<i>Admin qo'shish uchun .env faylini o'zgartiring</i>"
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:panel"))
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


# ============================================================
# EXCEL IMPORT
# ============================================================

@router.callback_query(F.data == "admin:import_excel")
async def import_excel_menu(callback: CallbackQuery):
    """Excel import instructions"""
    if not is_admin(callback.from_user.id):
        return
    
    text = """
📥 <b>Excel'dan import</b>

Excel faylni quyidagi formatda tayyorlang:

<b>Ustunlar:</b>
1. question - Savol matni
2. correct - To'g'ri javob
3. wrong1 - Xato variant 1
4. wrong2 - Xato variant 2
5. wrong3 - Xato variant 3
6. explanation - Tushuntirish (ixtiyoriy)

Keyin faylni yuboring va /import [day_id] buyrug'ini yozing.

<i>Masalan: /import 1</i>
"""
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="📄 Namuna yuklab olish", callback_data="admin:download_template"))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:questions"))
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.callback_query(F.data == "admin:download_template")
async def download_template(callback: CallbackQuery):
    """Download Excel template for questions import"""
    if not is_admin(callback.from_user.id):
        return
    
    await callback.answer("📄 Template tayyorlanmoqda...")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import tempfile
        import os
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Questions"
        
        headers = ["question", "correct", "wrong1", "wrong2", "wrong3", "explanation"]
        descriptions = ["Savol matni", "To'g'ri javob", "Xato 1", "Xato 2", "Xato 3", "Tushuntirish"]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, (h, d) in enumerate(zip(headers, descriptions), 1):
            ws.cell(row=1, column=col, value=h).font = header_font
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).border = border
            ws.cell(row=2, column=col, value=d).font = Font(italic=True, color="666666")
            ws.cell(row=2, column=col).border = border
        
        examples = [
            ["Wie heißen Sie?", "Ich heiße Anna", "Ich bin 20", "Aus Berlin", "Wohne hier", "Ism so'rash"],
            ["Gegenteil von 'groß'?", "klein", "lang", "breit", "hoch", "groß=katta, klein=kichik"],
        ]
        
        for r, ex in enumerate(examples, 3):
            for c, v in enumerate(ex, 1):
                ws.cell(row=r, column=c, value=v).border = border
        
        for col, w in enumerate([40, 25, 20, 20, 20, 35], 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        
        file_path = os.path.join(tempfile.gettempdir(), "questions_template.xlsx")
        wb.save(file_path)
        
        from aiogram.types import FSInputFile
        await callback.message.answer_document(
            document=FSInputFile(file_path, filename="questions_template.xlsx"),
            caption="📄 <b>Excel template</b>\n\nTo'ldiring va <code>/import [day_id]</code> bilan yuklang."
        )
        os.remove(file_path)
        
    except ImportError:
        await callback.message.answer("❌ <code>pip install openpyxl</code> kerak!")
    except Exception as e:
        await callback.message.answer(f"❌ Xatolik: {e}")

# ============================================================
# EXCEL IMPORT HANDLERS
# ============================================================

@router.message(F.document)
async def handle_document(message: Message, state: FSMContext):
    """Handle uploaded Excel files"""
    if not is_admin(message.from_user.id):
        return
    
    document = message.document
    if not document.file_name.endswith(('.xlsx', '.xls')):
        return
    
    await state.update_data(
        pending_excel_file_id=document.file_id,
        pending_excel_file_name=document.file_name
    )
    
    await message.answer(
        f"📄 <b>Fayl qabul qilindi:</b> {document.file_name}\n\n"
        "Import qilish uchun:\n"
        "<code>/import [day_id]</code>\n\n"
        "<i>Masalan: /import 1</i>"
    )

@router.message(Command("import"))
async def import_excel_command(message: Message, state: FSMContext, bot: Bot):
    """Import questions from Excel file"""
    if not is_admin(message.from_user.id):
        return
    
    args = message.text.split()
    if len(args) < 2:
        await message.answer("❌ <code>/import [day_id]</code>\nMasalan: <code>/import 1</code>")
        return
    
    try:
        day_id = int(args[1])
    except ValueError:
        await message.answer("❌ Day ID raqam bo'lishi kerak!")
        return
    
    data = await state.get_data()
    file_id = data.get("pending_excel_file_id")
    
    if not file_id:
        await message.answer("❌ Avval Excel faylni yuboring!")
        return
    
    await message.answer("⏳ Import qilinmoqda...")
    
    try:
        import openpyxl
        import tempfile
        import os
        
        file = await bot.get_file(file_id)
        file_path = os.path.join(tempfile.gettempdir(), f"import_{message.from_user.id}.xlsx")
        await bot.download_file(file.file_path, file_path)
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        required = ['question', 'correct', 'wrong1', 'wrong2', 'wrong3']
        missing = [c for c in required if c not in headers]
        
        if missing:
            await message.answer(f"❌ Ustunlar yo'q: {', '.join(missing)}")
            os.remove(file_path)
            return
        
        questions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            if not row_data.get('question') or row_data.get('question') == 'Savol matni':
                continue
            questions.append(row_data)
        
        if not questions:
            await message.answer("❌ Savollar topilmadi!")
            os.remove(file_path)
            return
        
        async with get_session() as session:
            repo = QuestionRepository(session)
            count = 0
            for q in questions:
                try:
                    await repo.create(
                        day_id=day_id,
                        question_text=q['question'],
                        option_a=q['correct'],
                        option_b=q['wrong1'],
                        option_c=q['wrong2'],
                        option_d=q['wrong3'],
                        correct_option="A",
                        explanation=q.get('explanation') or None
                    )
                    count += 1
                except Exception as e:
                    logger.error(f"Import error: {e}")
            await session.commit()
        os.remove(file_path)
        await state.update_data(pending_excel_file_id=None)
        
        await message.answer(
            f"✅ <b>Import yakunlandi!</b>\n\n"
            f"📊 Jami: {len(questions)} ta\n"
            f"✅ Yuklandi: {count} ta\n"
            f"📅 Day ID: {day_id}"
        )
        
    except ImportError:
        await message.answer("❌ <code>pip install openpyxl</code> kerak!")
    except Exception as e:
        logger.error(f"Import error: {e}")
        await message.answer(f"❌ Xatolik: {e}")


# ============================================================
# FLASHCARD EXCEL IMPORT
# ============================================================

@router.callback_query(F.data == "admin:flashcard_import")
async def flashcard_import_menu(callback: CallbackQuery):
    """Flashcard Excel import instructions"""
    if not is_admin(callback.from_user.id):
        return

    text = """
📥 <b>Flashcard Excel'dan import</b>

Excel faylni quyidagi formatda tayyorlang:

<b>Ustunlar:</b>
1. front - Old tomon (nemischa so'z)
2. back - Orqa tomon (o'zbekcha tarjima)
3. notes - Izoh (ixtiyoriy)
4. example - Misol gap (ixtiyoriy)

Keyin faylni yuboring va <code>/import_fc [deck_id]</code> buyrug'ini yozing.

<i>Masalan: /import_fc 1</i>
"""

    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="📄 Namuna yuklab olish", callback_data="admin:flashcard_template"))
    builder.row(InlineKeyboardButton(text="📋 Decklar ro'yxati", callback_data="admin:deck_list"))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:decks"))

    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.callback_query(F.data == "admin:flashcard_template")
async def download_flashcard_template(callback: CallbackQuery):
    """Download Excel template for flashcard import"""
    if not is_admin(callback.from_user.id):
        return

    await callback.answer("📄 Template tayyorlanmoqda...")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        import tempfile
        import os

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Flashcards"

        headers = ["front", "back", "notes", "example"]
        descriptions = ["Nemischa so'z", "O'zbekcha tarjima", "Izoh", "Misol gap"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, (h, d) in enumerate(zip(headers, descriptions), 1):
            ws.cell(row=1, column=col, value=h).font = header_font
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).border = border
            ws.cell(row=2, column=col, value=d).font = Font(italic=True, color="666666")
            ws.cell(row=2, column=col).border = border

        examples = [
            ["der Hund", "It", "Erkak jinsi", "Der Hund ist groß."],
            ["die Katze", "Mushuk", "Ayol jinsi", "Die Katze schläft."],
            ["das Kind", "Bola", "Neytral jinsi", "Das Kind spielt."],
            ["gehen", "Bormoq", "Fe'l", "Ich gehe nach Hause."],
            ["groß", "Katta", "Sifat", "Das Haus ist groß."],
        ]

        for r, ex in enumerate(examples, 3):
            for c, v in enumerate(ex, 1):
                ws.cell(row=r, column=c, value=v).border = border

        for col, w in enumerate([25, 25, 20, 40], 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        file_path = os.path.join(tempfile.gettempdir(), "flashcard_template.xlsx")
        wb.save(file_path)

        from aiogram.types import FSInputFile
        await callback.message.answer_document(
            document=FSInputFile(file_path, filename="flashcard_template.xlsx"),
            caption="📄 <b>Flashcard Excel template</b>\n\n"
                   "To'ldiring va <code>/import_fc [deck_id]</code> bilan yuklang.\n\n"
                   "<i>Deck ID ni ko'rish uchun: Decklar ro'yxati</i>"
        )
        os.remove(file_path)

    except ImportError:
        await callback.message.answer("❌ <code>pip install openpyxl</code> kerak!")
    except Exception as e:
        await callback.message.answer(f"❌ Xatolik: {e}")


@router.message(Command("import_fc"))
async def import_flashcard_command(message: Message, state: FSMContext, bot: Bot):
    """Import flashcards from Excel file"""
    if not is_admin(message.from_user.id):
        return

    args = message.text.split()
    if len(args) < 2:
        await message.answer(
            "❌ <code>/import_fc [deck_id]</code>\n"
            "Masalan: <code>/import_fc 1</code>\n\n"
            "Deck ID ni ko'rish uchun /admin → Decklar → Ro'yxat"
        )
        return

    try:
        deck_id = int(args[1])
    except ValueError:
        await message.answer("❌ Deck ID raqam bo'lishi kerak!")
        return

    data = await state.get_data()
    file_id = data.get("pending_excel_file_id")

    if not file_id:
        await message.answer("❌ Avval Excel faylni yuboring!")
        return

    # Check deck exists
    from src.database.models import FlashcardDeck, Flashcard
    from sqlalchemy import select

    async with get_session() as session:
        result = await session.execute(
            select(FlashcardDeck).where(FlashcardDeck.id == deck_id)
        )
        deck = result.scalar_one_or_none()

        if not deck:
            await message.answer(f"❌ Deck ID {deck_id} topilmadi!")
            return

    await message.answer(f"⏳ {deck.name} ga import qilinmoqda...")

    try:
        import openpyxl
        import tempfile
        import os

        file = await bot.get_file(file_id)
        file_path = os.path.join(tempfile.gettempdir(), f"import_fc_{message.from_user.id}.xlsx")
        await bot.download_file(file.file_path, file_path)

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        required = ['front', 'back']
        missing = [c for c in required if c not in headers]

        if missing:
            await message.answer(f"❌ Ustunlar yo'q: {', '.join(missing)}")
            os.remove(file_path)
            return

        cards = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            if not row_data.get('front') or row_data.get('front') in ['Nemischa so\'z', 'front']:
                continue
            cards.append(row_data)

        if not cards:
            await message.answer("❌ Kartalar topilmadi!")
            os.remove(file_path)
            return

        async with get_session() as session:
            count = 0
            for c in cards:
                try:
                    card = Flashcard(
                        deck_id=deck_id,
                        front_text=str(c['front']).strip(),
                        back_text=str(c['back']).strip(),
                        notes=str(c.get('notes') or '').strip() or None,
                        example_sentence=str(c.get('example') or '').strip() or None,
                        is_active=True
                    )
                    session.add(card)
                    count += 1
                except Exception as e:
                    logger.error(f"Flashcard import error: {e}")

            # Update deck cards count
            result = await session.execute(
                select(FlashcardDeck).where(FlashcardDeck.id == deck_id)
            )
            deck = result.scalar_one()

            # Count total cards
            from sqlalchemy import func
            count_result = await session.execute(
                select(func.count(Flashcard.id)).where(Flashcard.deck_id == deck_id)
            )
            total_cards = count_result.scalar()
            deck.cards_count = total_cards

            await session.commit()

        os.remove(file_path)
        await state.update_data(pending_excel_file_id=None)

        await message.answer(
            f"✅ <b>Import yakunlandi!</b>\n\n"
            f"📦 Deck: {deck.name}\n"
            f"📊 Yuklandi: {count} ta karta\n"
            f"📊 Jami deckda: {total_cards} ta karta"
        )

    except ImportError:
        await message.answer("❌ <code>pip install openpyxl</code> kerak!")
    except Exception as e:
        logger.error(f"Flashcard import error: {e}")
        await message.answer(f"❌ Xatolik: {e}")


# ============================================================
# FLASHCARD DECK MANAGEMENT
# ============================================================

def deck_menu_keyboard() -> InlineKeyboardMarkup:
    """Deck management menu"""
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="📋 Barcha decklar", callback_data="admin:deck_list"))
    builder.row(InlineKeyboardButton(text="➕ Yangi deck", callback_data="admin:deck_add"))
    builder.row(InlineKeyboardButton(text="📥 Excel'dan import", callback_data="admin:flashcard_import"))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:panel"))
    return builder.as_markup()


@router.callback_query(F.data == "admin:decks")
async def admin_decks_menu(callback: CallbackQuery):
    """Deck management menu"""
    if not is_admin(callback.from_user.id):
        await callback.answer("⛔ Ruxsat yo'q!", show_alert=True)
        return
    
    await callback.message.edit_text(
        "📦 <b>Flashcard Decklar</b>\n\n"
        "Bu yerda so'z kartalari to'plamlarini boshqarishingiz mumkin.",
        reply_markup=deck_menu_keyboard()
    )
    await callback.answer()


@router.callback_query(F.data == "admin:deck_list")
async def admin_deck_list(callback: CallbackQuery):
    """List all decks"""
    if not is_admin(callback.from_user.id):
        return
    
    from src.database.models import FlashcardDeck
    from sqlalchemy import select
    
    async with get_session() as session:
        result = await session.execute(
            select(FlashcardDeck).order_by(FlashcardDeck.id)
        )
        decks = result.scalars().all()
    
    if not decks:
        text = "📦 <b>Decklar</b>\n\n<i>Hozircha deck yo'q</i>"
    else:
        text = "📦 <b>Decklar ro'yxati:</b>\n\n"
        for d in decks:
            premium = "💎" if d.is_premium else "🆓"
            text += f"{d.icon} <b>{d.name}</b> {premium}\n"
            text += f"   📊 {d.cards_count} ta karta\n\n"
    
    builder = InlineKeyboardBuilder()
    for d in decks:
        builder.row(InlineKeyboardButton(
            text=f"{d.icon} {d.name} ({d.cards_count})",
            callback_data=f"admin:deck_view:{d.id}"
        ))
    builder.row(InlineKeyboardButton(text="➕ Yangi deck", callback_data="admin:deck_add"))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:decks"))
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.callback_query(F.data == "admin:deck_add")
async def admin_deck_add(callback: CallbackQuery, state: FSMContext):
    """Start adding new deck"""
    if not is_admin(callback.from_user.id):
        return
    
    await state.set_state(AdminStates.waiting_deck_name)
    await callback.message.edit_text(
        "📦 <b>Yangi deck qo'shish</b>\n\n"
        "Deck nomini kiriting:\n"
        "<i>Masalan: Tana a'zolari</i>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Bekor qilish", callback_data="admin:decks")]
        ])
    )
    await callback.answer()


@router.message(AdminStates.waiting_deck_name)
async def process_deck_name(message: Message, state: FSMContext):
    """Process deck name"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    
    # Edit mode
    if data.get("edit_deck_id"):
        from src.database.models import FlashcardDeck
        from sqlalchemy import select
        
        async with get_session() as session:
            result = await session.execute(
                select(FlashcardDeck).where(FlashcardDeck.id == data["edit_deck_id"])
            )
            deck = result.scalar_one_or_none()
            if deck:
                deck.name = message.text
                await session.commit()
        
        await state.clear()
        await message.answer(f"✅ Nom ozgartirildi: <b>{message.text}</b>")
        return
    
    # Create mode
    await state.update_data(deck_name=message.text)
    await state.set_state(AdminStates.waiting_deck_description)
    
    await message.answer(
        "📝 Deck tavsifini kiriting:\n"
        "<i>Masalan: Nemis tilida inson tanasi azolari</i>"
    )


@router.message(AdminStates.waiting_deck_description)
async def process_deck_description(message: Message, state: FSMContext):
    """Process deck description"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    
    # Edit mode
    if data.get("edit_deck_id"):
        from src.database.models import FlashcardDeck
        from sqlalchemy import select
        
        async with get_session() as session:
            result = await session.execute(
                select(FlashcardDeck).where(FlashcardDeck.id == data["edit_deck_id"])
            )
            deck = result.scalar_one_or_none()
            if deck:
                deck.description = message.text
                await session.commit()
        
        await state.clear()
        await message.answer(f"✅ Tavsif o\'zgartirildi!")
        return
    
    # Create mode
    await state.update_data(deck_description=message.text)
    await state.set_state(AdminStates.waiting_deck_icon)
    
    await message.answer(
        "🎨 Deck uchun emoji tanlang:\n"
        "<i>Masalan: 🧍 yoki 🍎 yoki 🏠</i>"
    )


@router.message(AdminStates.waiting_deck_icon)
async def process_deck_icon(message: Message, state: FSMContext):
    """Process deck icon"""
    if not is_admin(message.from_user.id):
        return
    
    icon = message.text.strip()[:10]
    data = await state.get_data()
    
    # Edit mode
    if data.get("edit_deck_id"):
        from src.database.models import FlashcardDeck
        from sqlalchemy import select
        
        async with get_session() as session:
            result = await session.execute(
                select(FlashcardDeck).where(FlashcardDeck.id == data["edit_deck_id"])
            )
            deck = result.scalar_one_or_none()
            if deck:
                deck.icon = icon
                await session.commit()
        
        await state.clear()
        await message.answer(f"✅ Emoji o\'zgartirildi: {icon}")
        return
    
    # Create mode
    await state.update_data(deck_icon=icon)
    await state.set_state(AdminStates.waiting_deck_price)
    
    await message.answer(
        "💰 Deck narxini kiriting (Stars):\n"
        "<i>0 = bepul, 50 = 50 stars</i>"
    )


@router.message(AdminStates.waiting_deck_price)
async def process_deck_price(message: Message, state: FSMContext):
    """Process deck price and create deck"""
    if not is_admin(message.from_user.id):
        return
    
    try:
        price = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Raqam kiriting!")
        return
    
    data = await state.get_data()
    
    from src.database.models import FlashcardDeck
    
    async with get_session() as session:
        deck = FlashcardDeck(
            name=data['deck_name'],
            description=data['deck_description'],
            icon=data['deck_icon'],
            is_premium=price > 0,
            is_public=True,
            cards_count=0
        )
        session.add(deck)
        await session.commit()
        await session.refresh(deck)
        deck_id = deck.id
    
    await state.clear()
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="🃏 Karta qo'shish", callback_data=f"admin:card_add:{deck_id}"))
    builder.row(InlineKeyboardButton(text="📋 Decklar", callback_data="admin:deck_list"))
    
    await message.answer(
        f"✅ <b>Deck yaratildi!</b>\n\n"
        f"{data['deck_icon']} <b>{data['deck_name']}</b>\n"
        f"📝 {data['deck_description']}\n"
        f"💰 {'Bepul' if price == 0 else f'{price} ⭐'}\n\n"
        f"Endi kartalar qo'shishingiz mumkin.",
        reply_markup=builder.as_markup()
    )


@router.callback_query(F.data.startswith("admin:deck_view:"))
async def admin_deck_view(callback: CallbackQuery):
    """View deck details"""
    if not is_admin(callback.from_user.id):
        return
    
    deck_id = int(callback.data.split(":")[-1])
    
    from src.database.models import FlashcardDeck, Flashcard
    from sqlalchemy import select
    
    async with get_session() as session:
        result = await session.execute(
            select(FlashcardDeck).where(FlashcardDeck.id == deck_id)
        )
        deck = result.scalar_one_or_none()
        
        if not deck:
            await callback.answer("❌ Deck topilmadi!", show_alert=True)
            return
        
        cards_result = await session.execute(
            select(Flashcard).where(Flashcard.deck_id == deck_id).limit(10)
        )
        cards = cards_result.scalars().all()
    
    premium = "💎 Premium" if deck.is_premium else "🆓 Bepul"
    text = f"{deck.icon} <b>{deck.name}</b> {premium}\n\n"
    text += f"📝 {deck.description or 'Tavsif yo`q'}\n"
    text += f"🃏 {deck.cards_count} ta karta\n\n"
    
    if cards:
        text += "<b>Kartalar:</b>\n"
        for c in cards:
            text += f"• {c.front_text} → {c.back_text}\n"
        if deck.cards_count > 10:
            text += f"<i>... va yana {deck.cards_count - 10} ta</i>\n"
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="🃏 Karta qo'shish", callback_data=f"admin:card_add:{deck_id}"))
    builder.row(
        InlineKeyboardButton(text="✏️ Tahrirlash", callback_data=f"admin:deck_edit:{deck_id}"),
        InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"admin:deck_del:{deck_id}")
    )
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:deck_list"))
    
    await callback.message.edit_text(text, reply_markup=builder.as_markup())
    await callback.answer()


@router.callback_query(F.data.startswith("admin:deck_del:"))
async def admin_deck_delete(callback: CallbackQuery):
    """Delete deck"""
    if not is_super_admin(callback.from_user.id):
        await callback.answer("⛔ Faqat super admin!", show_alert=True)
        return
    
    deck_id = int(callback.data.split(":")[-1])
    
    from src.database.models import FlashcardDeck
    from sqlalchemy import select, delete
    
    async with get_session() as session:
        await session.execute(delete(FlashcardDeck).where(FlashcardDeck.id == deck_id))
        await session.commit()
    
    await callback.answer("✅ Deck o'chirildi!", show_alert=True)
    await admin_deck_list(callback)


# ============================================================
# FLASHCARD CARD MANAGEMENT
# ============================================================

@router.callback_query(F.data == "admin:cards")
async def admin_cards_menu(callback: CallbackQuery):
    """Cards menu - select deck first"""
    if not is_admin(callback.from_user.id):
        return
    
    from src.database.models import FlashcardDeck
    from sqlalchemy import select
    
    async with get_session() as session:
        result = await session.execute(select(FlashcardDeck).order_by(FlashcardDeck.id))
        decks = result.scalars().all()
    
    if not decks:
        await callback.answer("❌ Avval deck yarating!", show_alert=True)
        return
    
    builder = InlineKeyboardBuilder()
    for d in decks:
        builder.row(InlineKeyboardButton(
            text=f"{d.icon} {d.name} ({d.cards_count})",
            callback_data=f"admin:deck_view:{d.id}"
        ))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data="admin:panel"))
    
    await callback.message.edit_text(
        "🃏 <b>Kartalar boshqaruvi</b>\n\n"
        "Karta qo'shish uchun deck tanlang:",
        reply_markup=builder.as_markup()
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin:card_add:"))
async def admin_card_add(callback: CallbackQuery, state: FSMContext):
    """Start adding card to deck"""
    if not is_admin(callback.from_user.id):
        return
    
    deck_id = int(callback.data.split(":")[-1])
    await state.update_data(deck_id=deck_id)
    await state.set_state(AdminStates.waiting_card_front)
    
    await callback.message.edit_text(
        "🃏 <b>Yangi karta qo'shish</b>\n\n"
        "Old tomoni (so'z/savol):\n"
        "<i>Masalan: der Kopf</i>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Bekor qilish", callback_data=f"admin:deck_view:{deck_id}")]
        ])
    )
    await callback.answer()


@router.message(AdminStates.waiting_card_front)
async def process_card_front(message: Message, state: FSMContext):
    """Process card front"""
    if not is_admin(message.from_user.id):
        return
    
    await state.update_data(card_front=message.text)
    await state.set_state(AdminStates.waiting_card_back)
    
    await message.answer(
        "🔙 Orqa tomoni (tarjima/javob):\n"
        "<i>Masalan: Bosh</i>"
    )


@router.message(AdminStates.waiting_card_back)
async def process_card_back(message: Message, state: FSMContext):
    """Process card back"""
    if not is_admin(message.from_user.id):
        return
    
    await state.update_data(card_back=message.text)
    await state.set_state(AdminStates.waiting_card_example)
    
    await message.answer(
        "📝 Misol gap (ixtiyoriy):\n"
        "<i>Masalan: Mein Kopf tut weh. - Boshim og'riyapti.</i>\n\n"
        "O'tkazib yuborish uchun <b>-</b> yozing"
    )


@router.message(AdminStates.waiting_card_example)
async def process_card_example(message: Message, state: FSMContext):
    """Process card example and save"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    example = message.text if message.text != "-" else None
    
    from src.database.models import FlashcardDeck, Flashcard
    from sqlalchemy import select
    
    async with get_session() as session:
        # Create card
        card = Flashcard(
            deck_id=data['deck_id'],
            front_text=data['card_front'],
            back_text=data['card_back'],
            example_sentence=example
        )
        session.add(card)
        
        # Update deck card count
        result = await session.execute(
            select(FlashcardDeck).where(FlashcardDeck.id == data['deck_id'])
        )
        deck = result.scalar_one()
        deck.cards_count += 1
        
        await session.commit()
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="➕ Yana qo'shish", callback_data=f"admin:card_add:{data['deck_id']}"))
    builder.row(InlineKeyboardButton(text="📋 Deckga qaytish", callback_data=f"admin:deck_view:{data['deck_id']}"))
    
    await state.clear()
    
    await message.answer(
        f"✅ <b>Karta qo'shildi!</b>\n\n"
        f"🔤 {data['card_front']}\n"
        f"🔙 {data['card_back']}\n"
        f"{'📝 ' + example if example else ''}",
        reply_markup=builder.as_markup()
    )


# ============================================================
# DECK EDIT HANDLER
# ============================================================

@router.callback_query(F.data.startswith("admin:deck_edit:"))
async def admin_deck_edit(callback: CallbackQuery):
    """Edit deck menu"""
    if not is_admin(callback.from_user.id):
        return
    
    deck_id = int(callback.data.split(":")[-1])
    
    from src.database.models import FlashcardDeck
    from sqlalchemy import select
    
    async with get_session() as session:
        result = await session.execute(
            select(FlashcardDeck).where(FlashcardDeck.id == deck_id)
        )
        deck = result.scalar_one_or_none()
    
    if not deck:
        await callback.answer("❌ Deck topilmadi!", show_alert=True)
        return
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="📝 Nomini o'zgartirish", callback_data=f"admin:deck_edit_name:{deck_id}"))
    builder.row(InlineKeyboardButton(text="📋 Tavsifni o'zgartirish", callback_data=f"admin:deck_edit_desc:{deck_id}"))
    builder.row(InlineKeyboardButton(text="🎨 Emoji o'zgartirish", callback_data=f"admin:deck_edit_icon:{deck_id}"))
    builder.row(InlineKeyboardButton(
        text=f"{'🆓 Bepul qilish' if deck.is_premium else '💎 Premium qilish'}",
        callback_data=f"admin:deck_toggle_premium:{deck_id}"
    ))
    builder.row(InlineKeyboardButton(text="◀️ Orqaga", callback_data=f"admin:deck_view:{deck_id}"))
    
    await callback.message.edit_text(
        f"✏️ <b>Deck tahrirlash</b>\n\n"
        f"{deck.icon} <b>{deck.name}</b>\n"
        f"📝 {deck.description or 'Tavsif yo`q'}\n"
        f"💎 {'Premium' if deck.is_premium else 'Bepul'}",
        reply_markup=builder.as_markup()
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin:deck_toggle_premium:"))
async def admin_deck_toggle_premium(callback: CallbackQuery):
    """Toggle deck premium status"""
    if not is_admin(callback.from_user.id):
        return
    
    deck_id = int(callback.data.split(":")[-1])
    
    from src.database.models import FlashcardDeck
    from sqlalchemy import select
    
    async with get_session() as session:
        result = await session.execute(
            select(FlashcardDeck).where(FlashcardDeck.id == deck_id)
        )
        deck = result.scalar_one_or_none()
        
        if deck:
            deck.is_premium = not deck.is_premium
            await session.commit()
            status = "💎 Premium" if deck.is_premium else "🆓 Bepul"
            await callback.answer(f"✅ Deck endi {status}", show_alert=True)
    
    # Refresh edit menu
    callback.data = f"admin:deck_edit:{deck_id}"
    await admin_deck_edit(callback)


@router.callback_query(F.data.startswith("admin:deck_edit_name:"))
async def admin_deck_edit_name_start(callback: CallbackQuery, state: FSMContext):
    """Start editing deck name"""
    if not is_admin(callback.from_user.id):
        return
    
    deck_id = int(callback.data.split(":")[-1])
    await state.update_data(edit_deck_id=deck_id, edit_field="name")
    await state.set_state(AdminStates.waiting_deck_name)
    
    await callback.message.edit_text(
        "📝 <b>Yangi nom kiriting:</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Bekor qilish", callback_data=f"admin:deck_edit:{deck_id}")]
        ])
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin:deck_edit_desc:"))
async def admin_deck_edit_desc_start(callback: CallbackQuery, state: FSMContext):
    """Start editing deck description"""
    if not is_admin(callback.from_user.id):
        return
    
    deck_id = int(callback.data.split(":")[-1])
    await state.update_data(edit_deck_id=deck_id, edit_field="desc")
    await state.set_state(AdminStates.waiting_deck_description)
    
    await callback.message.edit_text(
        "📋 <b>Yangi tavsif kiriting:</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Bekor qilish", callback_data=f"admin:deck_edit:{deck_id}")]
        ])
    )
    await callback.answer()


@router.callback_query(F.data.startswith("admin:deck_edit_icon:"))
async def admin_deck_edit_icon_start(callback: CallbackQuery, state: FSMContext):
    """Start editing deck icon"""
    if not is_admin(callback.from_user.id):
        return
    
    deck_id = int(callback.data.split(":")[-1])
    await state.update_data(edit_deck_id=deck_id, edit_field="icon")
    await state.set_state(AdminStates.waiting_deck_icon)
    
    await callback.message.edit_text(
        "🎨 <b>Yangi emoji kiriting:</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Bekor qilish", callback_data=f"admin:deck_edit:{deck_id}")]
        ])
    )
    await callback.answer()
